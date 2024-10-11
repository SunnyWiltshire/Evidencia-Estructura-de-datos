import csv
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font 
import pandas as pd
import numpy as np

unidades = {}
clientes = {}
prestamos = {}
ruta = []

def mostrar_ruta():
    print('\nRUTA: ')
    print(" > ".join(ruta))
#funcion que despliega el menu principal
def menu_principal():
    ruta.append('Menu Principal')
    while True:
        mostrar_ruta()
        print("\n--- MENÚ PRINCIPAL ---")
        print("1. Registro")
        print("2. Préstamo")
        print("3. Retorno")
        print("4. Informes")
        print("5. Salir\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append("Registro")
                menu_registro()
                ruta.pop()
            elif opcion == 2:
                ruta.append("Prestamo")
                registrar_prestamo(clientes, unidades, prestamos, rentas, conteo_rodadas)
                ruta.pop()
            elif opcion == 3:
                ruta.append("Retorno")
                menu_retorno()
                ruta.pop()
            elif opcion == 4:
                ruta.append("Informes")
                menu_informes()
                ruta.pop()
            elif opcion == 5:
                confirmacion = input("¿Desea salir del programa? (S/N)").upper()
                if confirmacion == "S":
                    print("Saliendo del sistema...\n")
                    break
                elif confirmacion == "N":
                    return
                else:
                    print("Opción invalida, ingrese los valores de 'S' o 'N'.")
            else:
                print("Opción invalida, intentalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numerico')

#funcion que pregunta al usuario si desea cancelar la accion que estaba haciendo
def cancelar():
    while True:
        try:
            respuesta = int(input("\nHa ocurrido un error. ¿Deseas cancelar o intentar de nuevo? \n1: cancelar  \n2: intentar de nuevo \n"))
            if respuesta == 1:
                print("Operacion cancelada.")
                return True
            elif respuesta == 2:
                return False
            else:
                print("Opción no valida. Por favor, selecciona 1 para cancelar o 2 para intentar de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numerico')
            
#funcion que despliega el sub menú de registro
def menu_registro():
    while True:
        mostrar_ruta()
        print("\n--- SUBMENÚ REGISTRO ---")
        print("1. Registrar una unidad")
        print("2. Registrar un cliente")
        print("3. Volver al menu principal\n")

        try:
            opcion = input("Elige una opción: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append('Unidad')
                registro_Unidad()
                ruta.pop()
            elif opcion == 2:
                ruta.append('Cliente')
                registro_Cliente()
                ruta.pop()
            elif opcion == 3:
                break
            else:
                print("Opción invalida, intentalo de nuevo.")
        except ValueError:
            if cancelar():
                break

## FUNCIONES PARA EL REGISTRO DE UNA UNIDAD

#funcion que permite registrar una unidad lista para un prestamo
def registro_Unidad():
    while True:
        mostrar_ruta()
        opcion = input("¿Deseas realizar un registro de unidad? (S/N): ").upper()

        if opcion == "S":
            print("\n--- REGISTRO DE UNIDAD ---")
            clave = max(unidades, default=0) + 1
            while True:
                entrada = input('Ingrese la rodada de la unidad (20, 26 o 29): ')
                rodada = int(entrada)
                if rodada in [20, 26, 29]:
                    break
                else:
                    print("Rodada no válida.")
                if cancelar():
                    return
            while True:
                print("""\nTenemos disponibles los siguientes colores: \nRojo \nAzul \nAmarillo \nVerde \nRosa""")
                color = input("Elige un color para la bicicleta: ").upper()
                if color in ["ROJO", "AZUL", "AMARILLO", "VERDE", "ROSA"]: 
                    print(f"Unidad registrada con exito. Clave: {clave}, Rodada: {rodada}, Color: {color}")
                    unidades[clave] = (entrada, color)
                    export_unidades_auto(unidades)
                    break
                else: 
                    print("Color no válido")
                if cancelar():
                    return
        elif opcion == "N":
            # Regresar al menú registro si elige 'N'
            break
        else:
            print("Opción inválida. Debes ingresar 'S' o 'N'.")
            if cancelar():
                break
            return

## Exporta automaticamente las unidades para su lectura
def export_unidades_auto(unidades):
    with open("Unidades_bicicletas.csv", "w", encoding="latin1", newline="") as archivocsv_unidades:
        grabador = csv.writer(archivocsv_unidades)
        grabador.writerow(("Clave", "Rodada", "Color"))
        grabador.writerows([(clave, datos[0], datos[1]) for clave, datos in unidades.items()])

## Lee las unidades para no perder los datos
def cargar_unidades_csv(nombre_archivo="Unidades_bicicletas.csv"):
    unidades = {}
    try:
        with open(nombre_archivo, "r", encoding="latin1", newline="") as archivocsv_unidades:
            lector = csv.reader(archivocsv_unidades)
            next(lector) 
            for fila in lector:
                clave, rodada, color = fila 
                unidades[int(clave)] = (int(rodada), color) 
    except FileNotFoundError:
        print("El archivo de unidades no existe. Se creará uno nuevo al exportar.")
    return unidades     

## FUNCIONES PARA EL REGISTRO DE UN CLIENTE

#funcion que permite registrar un cliente listo para solicitar un prestamo           
def registro_Cliente():
    while True:
        mostrar_ruta()
        opcion = input("¿Deseas realizar un registro de cliente? (S/N): ").upper()

        if opcion == "S":
            print("\n--- REGISTRO DE CLIENTE ---")
            clave_cliente = max(clientes, default=0) + 1
            # Captura de Apellidos
            while True:
                apellidos = input("Ingresa el apellido del cliente (max 40 caracteres): ")
                if apellidos.replace(" ", "").isalpha() and len(apellidos) <= 40:
                    break
                else:
                    print("Apellidos no válidos.")
                    if cancelar():
                        return
            # Captura de Nombre
            while True:
                nombre = input("Ingresa el nombre del cliente (max 40 caracteres): ")
                if nombre.replace(" ", "").isalpha() and len(nombre) <= 40:
                    break
                else:
                    print("Nombre no válido.")
                    if cancelar():
                        return
            # Captura de Teléfono
            while True:
                telefono = input("Ingrese el número de teléfono (10 dígitos): ")
                if telefono.isdigit() and len(telefono) == 10:
                    break
                else:
                    print("Teléfono no válido.")
                    if cancelar():
                        return
            # Registro del cliente en la base de datos
            clientes[clave_cliente] = (apellidos, nombre, telefono)
            print(f"Cliente registrado con éxito. Clave: {clave_cliente}, Nombre: {nombre} {apellidos}, Teléfono: {telefono}")
            # Llamada a función para exportar datos
            export_clientes_auto(clientes)
            # Salir del bucle después de exportar
            break
        elif opcion == "N":
            break  # Salir del bucle para regresar al menú
        else:
            print("Opción inválida. Debes ingresar 'S' o 'N'.")

## Exporta automaticamente los clientes para su lectura
def export_clientes_auto(clientes):
    with open("Clientes_bicicletas.csv", "w", encoding="latin1", newline="") as archivocsv_clientes:
        grabador = csv.writer(archivocsv_clientes)
        grabador.writerow(("Clave", "Apellidos", "Nombres", "Teléfono"))
        grabador.writerows([(clave, datos[0], datos[1], datos[2]) for clave, datos in clientes.items()])    

## Lee los clientes para no perder los datos
def cargar_clientes_csv(nombre_archivo="Clientes_bicicletas.csv"):
    clientes = {}
    try:
        with open(nombre_archivo, "r", encoding="latin1", newline="") as archivocsv_clientes:
            lector = csv.reader(archivocsv_clientes)
            next(lector)
            for fila in lector:
                clave, apellidos, nombres, telefono = fila
                clientes[int(clave)] = (apellidos, nombres, telefono)
    except FileNotFoundError:
        print("El archivo de clientes no existe. Se creará uno nuevo al exportar.")
    return clientes

## FUNCIONES PARA EL REGISTRO DE UN PRÉSTAMO

## Apartado para registrar los préstamos
def registrar_prestamo(clientes, unidades, prestamos, rentas, conteo_rodadas, conteo_colores):
    mostrar_ruta()
    while True:
        tab_prestamos(clientes, unidades)
        opcion = input("¿Deseas realizar un registro de préstamos? (S/N): ").upper()
        
        if opcion == "S":
            print("\n--- REGISTRO DE PRÉSTAMO ---")

            fecha_actual = datetime.now().date()
            folio = max(prestamos, default=0) + 1

            # Captura de la clave de la unidad
            while True:
                Clave_unidad = input("Clave de la unidad: ")
                if Clave_unidad.isdigit() and int(Clave_unidad) in unidades:
                    Clave_unidad = int(Clave_unidad)
                    break
                print("La clave de la unidad no es válida o no es un número.")
                if cancelar(): return

            # Captura de la clave del cliente
            while True:
                Clave_cliente = input("Clave del cliente: ")
                if Clave_cliente.isdigit() and int(Clave_cliente) in clientes:
                    Clave_cliente = int(Clave_cliente)
                    break
                print("La clave del cliente no es válida o no es un número.")
                if cancelar(): return

            # Elección de la fecha del préstamo
            while True:
                eleccion_de_fecha = input("¿Deseas que la fecha sea la del día de hoy?\n1. Sí\n2. No\nElige una opción: ")
                if eleccion_de_fecha.isdigit():
                    eleccion_de_fecha = int(eleccion_de_fecha)
                    if eleccion_de_fecha == 1:
                        fecha_prestamo = fecha_actual
                        break
                    elif eleccion_de_fecha == 2:
                        while True:
                            fecha_a_elegir = input("Indica la fecha del préstamo (MM/DD/AAAA): ")
                            try:
                                fecha_prestamo = datetime.strptime(fecha_a_elegir, "%m/%d/%Y").date()
                                if fecha_prestamo >= fecha_actual:
                                    break
                                print("La fecha no puede ser anterior a la actual.")
                            except ValueError:
                                print("Formato de fecha incorrecto, intenta de nuevo.")
                        break
                print("Opción inválida, intenta de nuevo.")
                if cancelar(): return

            # Cantidad de días del préstamo
            while True:
                Cantidad_de_dias = input("¿Cuántos días de préstamo solicitas?: ")
                if Cantidad_de_dias.isdigit() and int(Cantidad_de_dias) > 0:
                    Cantidad_de_dias = int(Cantidad_de_dias)
                    fecha_de_retorno = fecha_prestamo + timedelta(days=Cantidad_de_dias)
                    print(f"La fecha en la que se debe regresar la unidad es el: {fecha_de_retorno.strftime('%m/%d/%Y')}")
                    break
                print("La cantidad de días debe ser un número mayor a 0.")
                if cancelar(): return

            # Registro del préstamo
            prestamos[folio] = {
                'Clave_cliente': Clave_cliente,
                'Clave_unidad': Clave_unidad,
                'Fecha_prestamo': fecha_prestamo.strftime("%m/%d/%Y"),
                'Fecha_retorno': fecha_de_retorno.strftime('%m/%d/%Y'),
                'Cantidad_dias': Cantidad_de_dias,
                'Retorno': False
            }

            # Actualizar la cantidad de rentas
            if Clave_cliente in rentas:
                rentas[Clave_cliente] += 1
            else:
                rentas[Clave_cliente] = 1  # Inicializar si no existe

            # Actualizar conteo de rodadas
            rodada = int(unidades[Clave_unidad][0])  # Obtener la rodada de la unidad
            if rodada in conteo_rodadas:
                conteo_rodadas[rodada] += 1
            else:
                conteo_rodadas[rodada] = 1  # Inicializar si no existe

            # Actualizar conteo de colores
            color = unidades[Clave_unidad][1]  # Obtener el color de la unidad
            if color in conteo_colores:
                conteo_colores[color] += 1
            else:
                conteo_colores[color] = 1  # Inicializar si no existe

            # Guardar los préstamos y rentas
            export_prestamos_auto(prestamos)
            guardar_rentas_csv(rentas)
            export_conteo_rodada(conteo_rodadas)
            export_conteo_colores(conteo_colores)

            print(f"Préstamo registrado exitosamente. Folio: {folio}, Cliente: {clientes[Clave_cliente][1]} {clientes[Clave_cliente][0]}, Unidad: {Clave_unidad}, Fecha de Préstamo: {fecha_prestamo}")
            break
        if opcion == 'N':
            break
        else: 
            print("Favor de indicar un valor correcto (S/N)")
            if cancelar(): return

            
            
## Impresión tabular que muestra los clientes y unidades al momento de realizar un préstamo
def tab_prestamos(clientes, unidades):
    print(f"{'Clave del cliente':^15}{'Nombre del cliente':^41}{'Clave de la unidad':^20}{'Rodada':^10}")
    print("=" * 100)
    
    # Iterar sobre clientes y asociar unidades, si las claves coinciden
    for clave_cliente, datos_cliente in clientes.items():
        if clave_cliente in unidades:
            rodada, _ = unidades[clave_cliente]  # Descomponemos la tupla
            print(f"{clave_cliente:^15}{datos_cliente[1] + ' ' + datos_cliente[0]:^41}{clave_cliente:^20}{rodada:^10}")
        else:
            print(f"{clave_cliente:^15}{datos_cliente[1] + ' ' + datos_cliente[0]:^41}{'Sin unidad':^20}{'N/A':^10}")
    
    print("=" * 100)

## Exporta automaticamente los préstamos para su lectura
def export_prestamos_auto(prestamos):
    with open("Prestamos_bicicletas.csv", "w", encoding="latin1", newline="") as archivocsv_prestamo:
        grabador = csv.writer(archivocsv_prestamo)
        grabador.writerow(("Folio", "Clave Cliente", "Clave Unidad", "Fecha préstamo", "Fecha retorno", "Cantidad_días", "Retorno"))
        grabador.writerows([
            (
                folio,
                datos['Clave_cliente'],
                datos['Clave_unidad'],
                datos['Fecha_prestamo'],
                datos['Fecha_retorno'],
                datos['Cantidad_dias'],
                datos['Retorno']
            )
            for folio, datos in prestamos.items()
        ])

    
## Lee los préstamos para no perder los datos
def cargar_prestamos_csv(nombre_archivo="Prestamos_bicicletas.csv"):
    prestamos = {}
    try:
        with open(nombre_archivo, "r", encoding="latin1", newline="") as archivocsv_prestamos:
            lector = csv.reader(archivocsv_prestamos)
            # Saltar los encabezados
            next(lector)
            for fila in lector:
                # Verificar si la fila tiene el número correcto de columnas
                if len(fila) != 7:  # Ajusta este número si tienes más o menos columnas en tu CSV
                    continue
                
                folio, Clave_cliente, Clave_unidad, Fecha_prestamo, Fecha_de_retorno, Cantidad_dias, Retorno = fila
                
                # Asegurarse de que 'folio' es un número válido antes de convertir
                if folio.isdigit():
                    prestamos[int(folio)] = {
                        'Clave_cliente': Clave_cliente,
                        'Clave_unidad': Clave_unidad,
                        'Fecha_prestamo': Fecha_prestamo,
                        'Fecha_retorno': Fecha_de_retorno,
                        'Cantidad_dias': int(Cantidad_dias),  # Asegúrate de que sea un número
                        'Retorno': Retorno
                    }
    except FileNotFoundError:
        print("El archivo de préstamos no existe. Se creará uno nuevo al exportar.")
    
    return prestamos



## MENU DE RETORNO        
#Función que despliega menú para hacer el retorno de la unidad
def menu_retorno():
    mostrar_ruta()
    if prestamos:
      print("\n--- SUBMENÚ RETORNO ---")
      while True:
          opcion = input("¿Deseas retornar una unidad? \n 1. Si \n 2. No, volver al menu principal \n Elige una opción: \n")

          if opcion == "1":
                while True:
                  numdefolio = input("\nIngrese el número de folio de su préstamo: \n")
                  try:
                      numdefolio = int(numdefolio)
                      if numdefolio in prestamos:
                          today = datetime.now().date()
                          prestamos[numdefolio]["Retorno"] = True  #v2
                          print("Retornó su unidad exitosamente el día", today.strftime('%m/%d/%Y'), "\n")
                          export_prestamos_auto(prestamos)
                          break
                      else:
                          print("El número de folio no existe. Por favor, inténtalo de nuevo.")
                          if cancelar():
                            break
                  except ValueError:
                      print("Por favor, ingrese un número entero.")
                      if cancelar():
                        break
                break
          elif opcion == "2":
              break
    else:
      print("No hay ningún prestamo realizado.")

## MENU INFORMES
def menu_informes():
    mostrar_ruta()
    while True:
        print("\n--- MENÚ INFORMES ---")
        print("1. Reportes")
        print("2. Análisis")
        print("3. Volver al menú\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append("Reportes")
                submenu_reportes()
                ruta.pop()
            elif opcion == 2:
                ruta.append("Análisis")
                submenu_analisis()
                ruta.pop()
            elif opcion == 3:
                return False
            else:
                print("Opción invalida, intentalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numerico')

## MENU DE REPORTES
def submenu_reportes():

  while True:
    mostrar_ruta()
    print("\n--- SUBMENÚ REPORTES ---")
    print("1. Clientes.")
    print("2. Listado de unidades.")
    print("3. Retrasos.")
    print("4. préstamos por retornar.")
    print("5. préstamos por periodo.")
    print("6. Salir al menú principal\n")

    try:
        reporte_opcion = int(input("Elige alguna de las opciones mencionadas: "))
        if reporte_opcion == 1:
            ruta.append('Clientes')
            exportar_clientes()
            ruta.pop()
        elif reporte_opcion == 2:
            ruta.append('Unidades')
            exportar_unidades()
            ruta.pop()
        elif reporte_opcion == 3:
            ruta.append('Retrasos')
            retrasos()
            ruta.pop()
        elif reporte_opcion == 4:
            ruta.append('Prestamos por retornar')
            reporte_prestamos_por_retornar(prestamos)
            ruta.pop()
        elif reporte_opcion == 5:
            ruta.append('Prestamos por periodo')
            prestamos_por_periodo(prestamos)
            ruta.pop()
        elif reporte_opcion == 6:
            return False    
        else:
            print("Ingresa una opción válida")
    except Exception as error_name:
        print(f"Ha ocurrido un error: {error_name}")
        if cancelar():
            break

## SUBMENU REPORTES CLIENTES
def exportar_clientes():
    mostrar_ruta()
    while True:
        if clientes:
            tab_clientes(clientes)
            try:
                export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n4. Salir al submenú\n"))
                if export_opcion == 1:
                    export_csv_clientes(clientes)
                elif export_opcion == 2:
                    export_excel_clientes(clientes)
                elif export_opcion == 3:
                    export_csv_clientes(clientes)
                    export_excel_clientes(clientes)
                elif export_opcion == 4:
                    break
                else:
                    print("Elige una opcion valida")
                    if cancelar():
                        break
            except ValueError:
                    print("Error: Debes ingresar un número entero que sea válido.")
                    if cancelar():
                        break
            except Exception as name_error:
                    print(f"Ha ocurrido un error inesperado: {name_error}")
                    if cancelar():
                        break
            else:
                print("No hay clientes para exportar")
                break
def exportar_unidades():
    mostrar_ruta()
    while True:
        if unidades:
            tab_unidades(unidades)
            try:
                export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n4. Salir al submenú\n"))
                if export_opcion == 1:
                    export_csv_unidades()
                elif export_opcion == 2:
                    export_excel_unidades()
                elif export_opcion == 3:
                    export_csv_unidades()
                    export_excel_unidades()
                elif export_opcion == 4:
                    break
                else:
                    print("Elige una opcion valida")
                    if cancelar():
                        break
            except ValueError:
                    print("Error: Debes ingresar un número entero que sea válido.")
                    if cancelar():
                        break
            except Exception as name_error:
                    print(f"Ha ocurrido un error inesperado: {name_error}")
                    if cancelar():
                        break
            else:
                print("No hay clientes para exportar")
                break

def export_excel_unidades(unidades, name_excel="Unidades.xlsx"):
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Unidades"

    hoja["A1"].value = "Clave"
    hoja["B1"].value = "Rodada"
    hoja["C1"].value = "Color"

    
    hoja["A1"].font = Font(bold=True)
    hoja["B1"].font = Font(bold=True)
    hoja["C1"].font = Font(bold=True)

    i = 2

    for clave, (rodada, color) in clientes.items():
        hoja.cell(row=i, column=1).value = clave
        hoja.cell(row=i, column=2).value = rodada
        hoja.cell(row=i, column=3).value = color
        i += 1
        
    ajustar_ancho_columnas(hoja)
    libro.save(name_excel)
    print(f"Datos exportados con éxito en {name_excel}")

def tab_unidades(unidades):
    print(f"{'Clave':^8}{'Rodada': <10}{'Color'}")
    print("=" * 30)
    for clave, datos in unidades.items():
        print(f"{clave:^8}{datos[0]: <10}{datos[1]}")
    print("=" * 30)

def export_csv_unidades(unidades):
    with open("Unidades_bicicletas.csv", "w", encoding="latin1", newline="") as archivocsv_unidades:
        grabador = csv.writer(archivocsv_unidades)
        grabador.writerow(("Clave", "Rodada", "Color"))
        grabador.writerows([(clave, datos[0], datos[1]) for clave, datos in unidades.items()])
        print("Datos exportados con éxito en Unidades_bicicletas.csv")

## Impresión tabular que muestra los clientes
def tab_clientes(clientes):
    print(f"{'Clave':^8}{'Apellidos': <41}{'Nombres': <41}{'Teléfono'}")
    print("=" * 100)
    for clave, datos in clientes.items():
        print(f"{clave:^8}{datos[0]: <41}{datos[1]: <41}{datos[2]}")
    print("=" * 100)

## Exporta los clientes en formato csv
def export_csv_clientes(clientes):
    with open("Clientes_bicicletas.csv", "w", encoding="latin1", newline="") as archivocsv_clientes:
        grabador = csv.writer(archivocsv_clientes)
        grabador.writerow(("Clave", "Apellidos", "Nombres", "Teléfono"))
        grabador.writerows([(clave, datos[0], datos[1], datos[2]) for clave, datos in clientes.items()])
    print("Datos exportados con éxito en Clientes_bicicletas.csv")

## funcion para ajustar el ancho de las columnas en excel
def ajustar_ancho_columnas(hoja):
    for column_cells in hoja.columns:
        max_length = 0
        column = column_cells[0].column_letter  
        for cell in column_cells:
            try:  
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  
        hoja.column_dimensions[column].width = adjusted_width

## Exporta los clientes en formato excel
def export_excel_clientes(clientes, name_excel="Clientes.xlsx"):
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Clientes"

    hoja["A1"].value = "Clave"
    hoja["B1"].value = "Apellidos"
    hoja["C1"].value = "Nombres"
    hoja["D1"].value = "Teléfono"
    
    hoja["A1"].font = Font(bold=True)
    hoja["B1"].font = Font(bold=True)
    hoja["C1"].font = Font(bold=True)
    hoja["D1"].font = Font(bold=True)
    hoja["E1"].font = Font(bold=True)

    i = 2

    for clave, (apellidos, nombres, telefono) in clientes.items():
        hoja.cell(row=i, column=1).value = clave
        hoja.cell(row=i, column=2).value = apellidos
        hoja.cell(row=i, column=3).value = nombres
        hoja.cell(row=i, column=4).value = telefono
        i += 1
        
    ajustar_ancho_columnas(hoja)
    libro.save(name_excel)
    print(f"Datos exportados con éxito en {name_excel}")

## SUBMENU LISTADO DE UNIDADES
def listado_unidades():
    while True:
        print("\n--- LISTADO DE UNIDADES ---")
        print("1. Completo")
        print("2. Por rodada")
        print("3. Por color")
        print("4. Volver al menú de listado de unidades\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                analisis_completo()
            elif opcion == 2:
                analisis_rodada()
            elif opcion == 3:
                analisis_color()
            elif opcion == 4:
                return False
            else:
                print("Opción invalida, intentalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numerico')

## SUBMENU RETRASOS
def retrasos():
    print('print pa q no de error, favor d borrar')

## SUBMENU ANÁLISIS COMPLETO
def analisis_completo():
    print('lolol')

## SUBMENU ANÁLISIS POR RODADA
def analisis_rodada():
    print('lolol')

## SUBMENU ANÁLISIS POR COLOR
def analisis_color():
    print('lolol')

## SUBMENU REPORTES PRÉSTAMOS POR RETORNAR
def reporte_prestamos_por_retornar(prestamos):
    if prestamos:
        while True:
            mostrar_ruta()
            try:
                fecha_inicial = input("\nIngresa la fecha inicial (MM/DD/AAAA): ")
                fecha_inicial = datetime.strptime(fecha_inicial, "%m/%d/%Y").date()
                break
            except ValueError:
                print("Formato de fecha incorrecto, intenta de nuevo.")
                if cancelar():
                    break

        
        while True:
            try:
                fecha_final = input("Ingresa la fecha final (MM/DD/AAAA): ")
                fecha_final = datetime.strptime(fecha_final, "%m/%d/%Y").date()
                if fecha_final >= fecha_inicial:
                    break
                else:
                    print("La fecha final debe ser posterior o igual a la fecha inicial.")
                    if cancelar():
                        break
            except ValueError:
                print("Formato de fecha incorrecto, intenta de nuevo.")
                if cancelar():
                    break
                
        print(f"{'Folio':^8}{'Clave del Cliente': <20}{'Clave de la Unidad': <20}{'Fecha Préstamo': <20}{'Fecha Retorno'}")
        print("=" * 80)

        for folio, datos in prestamos.items():
            if datos['Retorno'] == 'False':
                fecha_prestamo = datetime.strptime(datos["Fecha_prestamo"], "%m/%d/%Y").date()
                fecha_retorno = datetime.strptime(datos["Fecha_retorno"], "%m/%d/%Y").date()

                if fecha_inicial <= fecha_retorno <= fecha_final:
                    print(f"{folio:^8}{datos['Clave_cliente']: <20}{datos['Clave_unidad']: <20}{datos['Fecha_prestamo']: <20}{datos['Fecha_retorno']}")

        print("=" * 80)

        export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n 4.No deseo exportarlo"))
        if export_opcion == 1:
            export_csv_prestamos_retornar(prestamos, fecha_inicial, fecha_final)
        elif export_opcion == 2:
            export_excel_prestamos_retornar(prestamos, fecha_inicial, fecha_final)
        elif export_opcion == 3:
            export_csv_prestamos_retornar(prestamos, fecha_inicial, fecha_final)
            export_excel_prestamos_retornar(prestamos, fecha_inicial, fecha_final)
        elif export_opcion == 4:
            return False
        else:
            print("Elige una opción válida (1, 2, 3 o 4).")
            if cancelar():
                return False
    else:
        print("No se encontró ningún préstamo.")

## Exporta los préstamos por retornar en formato excel
def export_excel_prestamos_retornar(prestamos, fecha_prestamo, fecha_de_retorno, name_excel="Prestamos_por_retornar.xlsx"):
    libro = openpyxl.Workbook()

    hoja = libro.active
    hoja.title = "préstamos"

    hoja["A1"].value = "Folio"
    hoja["B1"].value = "Clave de la unidad"
    hoja["C1"].value = "Clave del cliente"
    hoja["D1"].value = "Fecha prestamo"
    hoja["E1"].value = "Fecha de retorno"
    
    hoja["A1"].font = Font(bold=True)
    hoja["B1"].font = Font(bold=True)
    hoja["C1"].font = Font(bold=True)
    hoja["D1"].font = Font(bold=True)
    hoja["E1"].font = Font(bold=True)

    i = 2
    for folio, datos in prestamos.items():
        if not datos["Retorno"]:
            fecha_retorno = datetime.strptime(datos["Fecha_retorno"], "%m/%d/%Y").date()
            if fecha_prestamo <= fecha_retorno <= fecha_de_retorno:
                hoja.cell(row=i, column=1).value = folio
                hoja.cell(row=i, column=2).value = datos["Clave_unidad"]
                hoja.cell(row=i, column=3).value = datos["Clave_cliente"]
                hoja.cell(row=i, column=4).value = datos["Fecha_prestamo"]
                hoja.cell(row=i, column=5).value = datos["Fecha_retorno"]
                i += 1
    ajustar_ancho_columnas(hoja)
    libro.save(name_excel)
    print(f"Datos exportados con éxito en {name_excel}")

## Exporta los préstamos por retornar en formato csv
def export_csv_prestamos_retornar(prestamos, fecha_prestamo, fecha_de_retorno, nombre_csv="Prestamos_por_retornar.csv"):
    with open(nombre_csv, "w", encoding="latin1", newline="") as archivo_csv:
        grabador = csv.writer(archivo_csv)

        grabador.writerow(("Folio", "Clave de la unidad", "Clave del cliente", "Fecha prestamo", "Fecha de retorno"))


        prestamos_filtrados = [
            (folio, datos["Clave_unidad"], datos["Clave_cliente"], datos["Fecha_prestamo"], datos["Fecha_retorno"])
            for folio, datos in prestamos.items()
            if fecha_prestamo <= datetime.strptime(datos['Fecha_prestamo'], "%m/%d/%Y").date() <= fecha_de_retorno
        ]

        if prestamos_filtrados:
            grabador.writerows(prestamos_filtrados)
            print(f"Datos exportados con éxito en {nombre_csv}")
        else:
            print("No hay préstamos que coincidan con los criterios especificados.")

## SUBMENU PRÉSTAMOS POR PERIODO
def prestamos_por_periodo():
    if prestamos:
            while True:
                mostrar_ruta()
                try:
                    fecha_inicial = input("\nIngresa la fecha inicial del periodo (MM/DD/AAAA): ")
                    fecha_inicial = datetime.strptime(fecha_inicial, "%m/%d/%Y").date()
                    break
                except ValueError:
                    print("Formato de fecha incorrecto, intenta de nuevo.")
                    if cancelar():
                        break

            while True:
                try:
                    fecha_final = input("Ingresa la fecha final del periodo (MM/DD/AAAA): ")
                    fecha_final = datetime.strptime(fecha_final, "%m/%d/%Y").date()
                    if fecha_final >= fecha_inicial:
                        break
                    else:
                        print("La fecha final debe ser posterior o igual a la fecha inicial.")
                        if cancelar():
                            break
                except ValueError:
                    print("Formato de fecha incorrecto, intenta de nuevo.")
                    if cancelar():
                        break

            print(f"{'Folio':^8}{'Clave del Cliente': <20}{'Clave de la Unidad': <20}{'Fecha Préstamo': <20}{'Fecha Retorno'}")
            print("=" * 80)

            for folio, datos in prestamos.items():
                fecha_prestamo = datetime.strptime(datos['Fecha_prestamo'], "%m/%d/%Y").date()
                if fecha_inicial <= fecha_prestamo <= fecha_final:
                    print(f"{folio:^8}{datos['Clave_cliente']: <20}{datos['Clave_unidad']: <20}{datos['Fecha_prestamo']: <20}{datos['Fecha_retorno']}")

            print("=" * 80)

            export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n"))
            if export_opcion == 1:
                export_csv_prestamos_por_periodo(prestamos, fecha_inicial, fecha_final)
            elif export_opcion == 2:
                export_excel_prestamos_por_periodo(prestamos, fecha_inicial, fecha_final)
            elif export_opcion == 3:
                export_csv_prestamos_por_periodo(prestamos, fecha_inicial, fecha_final)
                export_excel_prestamos_por_periodo(prestamos, fecha_inicial, fecha_final)
            elif export_opcion == 4:
                return False
            else:
                print("Elige una opción válida (1, 2, 3 o 4).")
                if cancelar():
                    return False
    else: 
        print("No hay préstamos para realizar un reporte")

## Exporta los préstamos por periodo en formato excel
def export_excel_prestamos_por_periodo(prestamos, fecha_prestamo, fecha_de_retorno, name_excel="Prestamos_por_periodo.xlsx"):
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "préstamos"

    hoja["A1"].value = "Folio"
    hoja["B1"].value = "Clave de la unidad"
    hoja["C1"].value = "Clave del cliente"
    hoja["D1"].value = "Fecha préstamo"
    hoja["E1"].value = "Fecha de retorno"
    
    hoja["A1"].font = Font(bold=True)
    hoja["B1"].font = Font(bold=True)
    hoja["C1"].font = Font(bold=True)
    hoja["D1"].font = Font(bold=True)
    hoja["E1"].font = Font(bold=True)

    # Poner negrita en los encabezados (fila 1)
    hoja["A1"].font = Font(bold=True)
    hoja["B1"].font = Font(bold=True)
    hoja["C1"].font = Font(bold=True)
    hoja["D1"].font = Font(bold=True)
    hoja["E1"].font = Font(bold=True)

    i = 2  # Iniciar en la fila 2 porque la fila 1 son los encabezados
    for folio, datos in prestamos.items():
        fecha_prestamo = datetime.strptime(datos['Fecha_prestamo'], "%m/%d/%Y").date()
        if fecha_prestamo <= fecha_prestamo <= fecha_de_retorno:
            # Asignar los valores a las celdas
            hoja.cell(row=i, column=1).value = folio
            hoja.cell(row=i, column=2).value = datos["Clave_unidad"]
            hoja.cell(row=i, column=3).value = datos["Clave_cliente"]
            hoja.cell(row=i, column=4).value = datos["Fecha_prestamo"]
            hoja.cell(row=i, column=5).value = datos["Fecha_retorno"]

            # Calcular el ancho según la longitud del contenido de cada fila
            hoja.column_dimensions["A"].width = max(hoja.column_dimensions["A"].width or 0, len(str(folio)))
            hoja.column_dimensions["B"].width = max(hoja.column_dimensions["B"].width or 0, len(str(datos["Clave_unidad"])))
            hoja.column_dimensions["C"].width = max(hoja.column_dimensions["C"].width or 0, len(str(datos["Clave_cliente"])))
            hoja.column_dimensions["D"].width = max(hoja.column_dimensions["D"].width or 0, len(str(datos["Fecha_prestamo"])))
            hoja.column_dimensions["E"].width = max(hoja.column_dimensions["E"].width or 0, len(str(datos["Fecha_retorno"])))

            i += 1  # Incrementa la fila
    ajustar_ancho_columnas(hoja)
    # Guarda el archivo Excel
    libro.save(name_excel)
    
    print(f"Datos exportados con éxito en {name_excel}")

## Exporta los préstamos por periodo en formato csv
def export_csv_prestamos_por_periodo(prestamos, fecha_prestamo, fecha_de_retorno, nombre_csv="Prestamos_por_periodo.csv"):
    with open(nombre_csv, "w", encoding="latin1", newline="") as archivo_csv:
        grabador = csv.writer(archivo_csv)
        grabador.writerow(("Folio", "Clave de la unidad", "Clave del cliente", "Fecha préstamo", "Fecha de retorno"))

        prestamos_filtrados = [
            (folio, datos["Clave_unidad"], datos["Clave_cliente"], datos["Fecha_prestamo"], datos["Fecha_retorno"])
            for folio, datos in prestamos.items()
            if fecha_prestamo <= datetime.strptime(datos['Fecha_prestamo'], "%m/%d/%Y").date() <= fecha_de_retorno
        ]

        if prestamos_filtrados:
            grabador.writerows(prestamos_filtrados)
            print(f"Datos exportados con éxito en {nombre_csv}")
        else:
            print("No hay préstamos que coincidan con los criterios especificados.")
            if cancelar():
                return False

##Submenú analísis
def submenu_analisis():
    mostrar_ruta()
    while True:
        print("\n--- SUBMENÚ ANÁLISIS ---")
        print("1. Duración de los préstamos.")
        print("2. Ranking de clientes.")
        print("3. Preferencias de rentas.")
        print("4. Volver al menú\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                duracion_prestamos(prestamos)
            elif opcion == 2:
                ranking_clientes(prestamos, clientes, rentas)
            elif opcion == 3:
                preferencias_rentas()
            elif opcion == 4:
                return False
            else:
                print("Opción invalida, intentalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numerico')
            
## SUBMENÚ DURACIÓN DE LOS PRÉSTAMOS
def duracion_prestamos(prestamos):
    dias_prestamo = [prestamo['Cantidad_dias'] for prestamo in prestamos.values()]

    if len(dias_prestamo) == 0:
        print("No hay registros de préstamos para calcular estadísticas.")
        return

    df = pd.DataFrame(dias_prestamo, columns=['Días de préstamo'])

    media = df['Días de préstamo'].mean()
    mediana = df['Días de préstamo'].median()
    moda = df['Días de préstamo'].mode().tolist()  # Convertir a lista en caso de múltiples modas
    minimo = df['Días de préstamo'].min()
    maximo = df['Días de préstamo'].max()
    desviacion_estandar = df['Días de préstamo'].std()
    cuartiles = np.percentile(df['Días de préstamo'], [25, 50, 75])

    reporte = {
        "Media": media,
        "Mediana": mediana,
        "Moda": moda,
        "Mínimo": minimo,
        "Máximo": maximo,
        "Desviación estándar": desviacion_estandar,
        "Cuartiles (25%, 50%, 75%)": cuartiles
    }

    for clave, valor in reporte.items():
        print(f"{clave}: {valor}")


#cargar rentas
def cargar_rentas_csv():
    rentas = {}
    try:
        with open('rentas.csv', mode='r') as file:
            reader = csv.reader(file)
            next(reader)  # Saltar la fila de encabezados
            for row in reader:
                clave_cliente = int(row[0])
                cantidad_rentas = int(row[1])
                rentas[clave_cliente] = cantidad_rentas
        print("Rentas cargadas correctamente.")
    except FileNotFoundError:
        print("Archivo de rentas no encontrado. Se inicializará una lista vacía.")
    except Exception as e:
        print(f"Error al cargar rentas: {e}")
    return rentas
    
#guardar rentas
def guardar_rentas_csv(rentas):
    try:
        df = pd.DataFrame(list(rentas.items()), columns=['Clave_cliente', 'Cantidad_rentas'])
        df.to_csv('rentas.csv', index=False)
    except Exception as e:
        print(f"Ocurrió un error al guardar las rentas: {e}")
    
## SUBMENÚ RANKING CLIENTES

# Función para generar el ranking de clientes
def ranking_clientes(prestamos, clientes, rentas):
    ranking_data = {
        'Cantidad_rentas': [],
        'Clave_cliente': [],
        'Nombre_completo': [],
        'Teléfono': []
    }

    # Contar las rentas acumuladas por cada cliente
    for clave_cliente, cantidad_rentas in rentas.items():
        if clave_cliente in clientes:
            cliente = clientes[clave_cliente]
            apellidos, nombre, telefono = cliente  # Desempaquetar la tupla

            ranking_data['Cantidad_rentas'].append(cantidad_rentas)
            ranking_data['Clave_cliente'].append(clave_cliente)
            ranking_data['Nombre_completo'].append(f"{nombre} {apellidos}")
            ranking_data['Teléfono'].append(telefono)

    # df para ordenar los resultados
    df_ranking = pd.DataFrame(ranking_data)
    df_ranking.sort_values(by='Cantidad_rentas', ascending=False, inplace=True)

    # ranking con formato de tabla
    print("\n--- RANKING DE CLIENTES ---")
    print(f"{'Posición':<10} {'Clave Cliente':<15} {'Nombre Completo':<30} {'Teléfono':<15} {'Cantidad de Rentas':<20}")
    print("=" * 90)
    for i, row in enumerate(df_ranking.itertuples(index=False), 1):
        print(f"{i:<10} {row.Clave_cliente:<15} {row.Nombre_completo:<30} {row.Teléfono:<15} {row.Cantidad_rentas:<20}")
    
    guardar_rentas_csv(rentas)
    guardar_ranking_csv(df_ranking)


def guardar_ranking_csv(df_ranking):
    df_ranking.to_csv("Ranking_clientes.csv", index=False, encoding="latin1")
    print("Ranking de clientes exportado exitosamente en 'Ranking_clientes.csv'.")



## SUBMENÚ PREFERENCIAS RENTAS
def preferencias_rentas():
    print("Elige el reporte que deseas generar:")
    print("1. Cantidad de préstamos por rodada")
    print("2. Cantidad de préstamos por color")
    
    while True:
        opcion_pref = input("Ingresa una opción (1 o 2): ")
        if opcion_pref.isdigit():
            opcion_pref = int(opcion_pref)
            if opcion_pref == 1:
                reporte_prestamos_por_rodada(conteo_rodadas)
                break
            elif opcion_pref == 2:
                reporte_prestamos_por_color(prestamos, unidades)
                break
            else:
                print("Opción inválida. Debes ingresar 1 o 2.")
        else:
            print("Entrada inválida. Por favor ingresa un número (1 o 2).")

def reporte_prestamos_por_rodada(conteo_rodadas):
    # Ordenar las rodadas por la cantidad de préstamos en orden descendente
    datos_ordenados = sorted(conteo_rodadas.items(), key=lambda x: x[1], reverse=True)

    # Imprimir el reporte en formato tabular
    print("\n--- REPORTE DE PRÉSTAMOS POR RODADA ---")
    print("{:<10} {:<20}".format("Rodada", "Cantidad de Préstamos"))
    print("-" * 30)
    for rodada, cantidad in datos_ordenados:
        print("{:<10} {:<20}".format(rodada, cantidad))




def export_conteo_rodada(conteo_rodadas, nombre_archivo="Conteo_Rodadas.csv"):
    # Exportar el conteo de rodadas a un archivo CSV
    with open(nombre_archivo, "w", encoding="latin1", newline="") as archivo_csv:
        grabador = csv.writer(archivo_csv)
        grabador.writerow(("Rodada", "Cantidad de Préstamos"))
        for rodada, cantidad in conteo_rodadas.items():
            grabador.writerow((rodada, cantidad))
    print(f"Conteo de rodadas exportado exitosamente en '{nombre_archivo}'")


def cargar_conteo_rodadas(nombre_archivo="Conteo_Rodadas.csv"):
    conteo_rodadas = {}
    try:
        # Abrir el archivo CSV para leer el conteo de rodadas
        with open(nombre_archivo, "r", encoding="latin1", newline="") as archivo_csv:
            lector = csv.reader(archivo_csv)
            # Saltar la fila de encabezado
            next(lector)
            # Leer cada fila y actualizar el diccionario conteo_rodadas
            for fila in lector:
                if len(fila) == 2:  # Asegurar que la fila tiene exactamente 2 columnas
                    rodada, cantidad = fila
                    conteo_rodadas[int(rodada)] = int(cantidad)
    except FileNotFoundError:
        print(f"El archivo '{nombre_archivo}' no existe. Asegúrate de que el archivo se haya exportado previamente.")
    
    return conteo_rodadas
    
    
   def reporte_prestamos_por_color(prestamos, unidades):
    # Crear un diccionario para contar los préstamos por color
    conteo_color = {}

    # Recorrer los préstamos para contar los colores
    for prestamo in prestamos.values():
        clave_unidad = prestamo['Clave_unidad']
        if clave_unidad in unidades:
            color = unidades[clave_unidad][1]  # Suponiendo que el color está en el segundo elemento de la tupla
            if color in conteo_color:
                conteo_color[color] += 1
            else:
                conteo_color[color] = 1

    # Convertir los datos en una lista de tuplas y ordenarlos por cantidad de préstamos (descendente)
    datos_ordenados = sorted(conteo_color.items(), key=lambda x: x[1], reverse=True)

    # Imprimir el reporte en formato tabular
    print("\n--- REPORTE DE PRÉSTAMOS POR COLOR ---")
    print("{:<15} {:<20}".format("Color", "Cantidad de Préstamos"))
    print("-" * 35)
    for color, cantidad in datos_ordenados:
        print("{:<15} {:<20}".format(color, cantidad))

def cargar_conteo_colores(nombre_archivo="Conteo_Colores.csv"):
    conteo_colores = {}
    try:
        # Abrir el archivo CSV para leer el conteo de colores
        with open(nombre_archivo, "r", encoding="latin1", newline="") as archivo_csv:
            lector = csv.reader(archivo_csv)
            # Saltar la fila de encabezado
            next(lector)
            # Leer cada fila y actualizar el diccionario conteo_colores
            for fila in lector:
                if len(fila) == 2:  # Asegurarse de que la fila tiene exactamente 2 columnas
                    color, cantidad = fila
                    conteo_colores[color] = int(cantidad)
    except FileNotFoundError:
        print(f"El archivo '{nombre_archivo}' no existe. Asegúrate de que el archivo se haya exportado previamente.")
    
    return conteo_colores 
    
    
    
# Inicio del programa
conteo_colores = cargar_conteo_colores()
conteo_rodadas = cargar_conteo_rodadas()
rentas = cargar_rentas_csv()
clientes = cargar_clientes_csv()
unidades = cargar_unidades_csv()
prestamos = cargar_prestamos_csv()
print("===== BIENVENIDO A NUESTRA RENTA DE BICICLETAS =====")
menu_principal()
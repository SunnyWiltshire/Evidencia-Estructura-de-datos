import csv
import datetime
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd
import sqlite3
from sqlite3 import Error
from tabulate import tabulate
from datetime import datetime, timedelta
import sys

ruta = []
try:
    with sqlite3.connect('RentaBicicletas.db') as conn: 
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS UNIDAD \
            (Clave INTEGER NOT NULL PRIMARY KEY, \
            Rodada INTEGER NOT NULL, \
            Color TEXT NOT NULL);")
        cursor.execute("CREATE TABLE IF NOT EXISTS CLIENTES \
            (Clave INTEGER NOT NULL PRIMARY KEY, \
            Apellidos TEXT NOT NULL, \
            Nombres TEXT NOT NULL, \
            Telefono INTEGER NOT NULL);")
        cursor.execute("CREATE TABLE IF NOT EXISTS PRESTAMO \
            (Folio INTEGER NOT NULL PRIMARY KEY, \
            Fecha_Prestamo INTEGER NOT NULL, \
            Dias_Prestamo INTEGER NOT NULL, \
            Fecha_Retorno INTEGER NOT NULL, \
            Retorno INTEGER NOT NULL, \
            Clave_Cliente INTEGER NOT NULL, \
            Clave_Unidad INTEGER NOT NULL, \
            FOREIGN KEY (Clave_Cliente) REFERENCES CLIENTES(Clave), \
            FOREIGN KEY (Clave_Unidad) REFERENCES UNIDAD(Clave));")
        print("Base de datos y tablas creadas exitosamente.") 
except Error as e:
    print(e)
except Exception:
    print(f"Se produjo el siguiente error: {sys.exc_info()}")
    
def mostrar_ruta():
    print('\nRUTA: ')
    print(" > ".join(ruta))
    
#funcion que despliega el menu principal
def menu_principal():
    ruta.append('Menú Principal')
    while True:
        mostrar_ruta()
        print("\n--- MENÚ PRINCIPAL ---")
        print("1. Registro")
        print("2. Préstamo")
        print("3. Retorno")
        print("4. Informes")
        print("5. Salir\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append("Registro")
                menu_registro()
                ruta.pop()
            elif opcion == 2:
                ruta.append("Préstamo")
                registrar_prestamo()
                ruta.pop()
            elif opcion == 3:
                ruta.append("Retorno")
                menu_retorno()
                ruta.pop()
            elif opcion == 4:
                ruta.append("Informes")
                menu_informes()
                ruta.pop()
            elif opcion == 5:
                confirmacion = input("¿Desea salir del programa? (S/N): ").upper()
                if confirmacion == "S":
                    print("Saliendo del sistema...\n")
                    break
                elif confirmacion == "N":
                    print("Continuando en el menú principal...")
                else:
                    print("Opción inválida, ingrese los valores de 'S' o 'N'.")
            else:
                print("Opción inválida, inténtalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numérico')


#funcion que pregunta al usuario si desea cancelar la accion que estaba haciendo
def cancelar():
    while True:
        try:
            respuesta = int(input("\nHa ocurrido un error. ¿Deseas cancelar o intentar de nuevo? \n1: cancelar  \n2: intentar de nuevo \n"))
            if respuesta == 1:
                print("Operación cancelada.")
                return True
            elif respuesta == 2:
                return False
            else:
                print("Opción no válida. Por favor, selecciona 1 para cancelar o 2 para intentar de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numérico')
            
#funcion que despliega el sub menú de registro
def menu_registro():
    while True:
        mostrar_ruta()
        print("\n--- SUBMENÚ REGISTRO ---")
        print("1. Registrar una unidad")
        print("2. Registrar un cliente")
        print("3. Volver al menú principal\n")

        try:
            opcion = input("Elige una opción: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append('Unidad')
                registro_Unidad()
                ruta.pop()
            elif opcion == 2:
                ruta.append('Cliente')
                registro_Cliente()
                ruta.pop()
            elif opcion == 3:
                break
            else:
                print("Opción inválida, inténtalo de nuevo.")
        except ValueError:
            if cancelar():
                break

## FUNCIONES PARA EL REGISTRO DE UNA UNIDAD

#funcion que permite registrar una unidad lista para un prestamo
def registro_Unidad():
    try:
        # Conectar a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            while True:
                mostrar_ruta()
                opcion = input("¿Deseas realizar un registro de unidad? (S/N): ").upper()

                if opcion == "S":
                    print("\n--- REGISTRO DE UNIDAD ---")
                    
                    # Obtener la clave más alta existente en la tabla UNIDAD y sumar 1
                    cursor.execute("SELECT IFNULL(MAX(Clave), 0) FROM UNIDAD")
                    clave = cursor.fetchone()[0] + 1
                    
                    while True:
                        entrada = input('Ingrese la rodada de la unidad (20, 26 o 29): ')
                        try:
                            rodada = int(entrada)
                            if rodada in [20, 26, 29]:
                                print("""\nTenemos disponibles los siguientes colores: \nRojo \nAzul \nAmarillo \nVerde \nRosa""")
                                color = input("Elige un color para la bicicleta: ").upper()
                                
                                if color in ["ROJO", "AZUL", "AMARILLO", "VERDE", "ROSA"]:
                                    # Insertar los datos en la tabla UNIDAD
                                    cursor.execute("INSERT INTO UNIDAD (Clave, Rodada, Color) VALUES (?, ?, ?)", 
                                                   (clave, rodada, color))
                                    conn.commit() 

                                    print(f"Unidad registrada con éxito. Clave: {clave}, Rodada: {rodada}, Color: {color}")
                                    return False
                                else:
                                    print("Color no válido. Por favor, elija entre Rojo, Azul, Amarillo, Verde, o Rosa.")
                                    if cancelar():
                                        return
                            else:
                                print("Por favor, ingrese un valor válido (20, 26 o 29).")
                                if cancelar():
                                    break

                        except ValueError:
                            print('Favor de ingresar un valor válido\n')
                            if cancelar():
                                break
                elif opcion == "N":
                    return False
                else:
                    print("Opción inválida. Debes ingresar 'S' o 'N'.")
                    if cancelar():
                        break
                    return
    except Error as e:
        print(f"Error de base de datos: {e}")
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()}")

## FUNCIONES PARA EL REGISTRO DE UN CLIENTE

#funcion que permite registrar un cliente          
def registro_Cliente():
    try:
        # Conectar a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            while True:
                mostrar_ruta()
                opcion = input("¿Deseas realizar un registro de cliente? (S/N): ").upper()

                if opcion == "S":
                    print("\n--- REGISTRO DE CLIENTE ---")
                    
                    # Obtener la clave más alta existente en la tabla CLIENTES y sumar 1
                    cursor.execute("SELECT IFNULL(MAX(Clave), 0) FROM CLIENTES")
                    clave_cliente = cursor.fetchone()[0] + 1
                    
                    # Captura de Apellidos
                    while True:
                        apellidos = input("Ingresa el apellido del cliente (max 40 caracteres): ")
                        if apellidos.replace(" ", "").isalpha() and len(apellidos) <= 40:
                            break
                        else:
                            print("Apellidos no válidos.")
                            if cancelar():
                                return
                    
                    # Captura de Nombre
                    while True:
                        nombre = input("Ingresa el nombre del cliente (max 40 caracteres): ")
                        if nombre.replace(" ", "").isalpha() and len(nombre) <= 40:
                            break
                        else:
                            print("Nombre no válido.")
                            if cancelar():
                                return
                    
                    # Captura de Teléfono
                    while True:
                        telefono = input("Ingrese el número de teléfono (10 dígitos): ")
                        if telefono.isdigit() and len(telefono) == 10:
                            break
                        else:
                            print("Teléfono no válido.")
                            if cancelar():
                                return
                    
                    # Insertar el cliente en la tabla CLIENTES
                    cursor.execute("INSERT INTO CLIENTES (Clave, Apellidos, Nombres, Telefono) VALUES (?, ?, ?, ?)", 
                                   (clave_cliente, apellidos, nombre, telefono))
                    conn.commit()  

                    print(f"Cliente registrado con éxito. Clave: {clave_cliente}, Nombre: {nombre} {apellidos}, Teléfono: {telefono}")
                    
                    # Salir del bucle después de registrar
                    break
                elif opcion == "N":
                    break  # Salir del bucle para regresar al menú
                else:
                    print("Opción inválida. Debes ingresar 'S' o 'N'.")
                    if cancelar():
                        break
                    return
    except Error as e:
        print(f"Error de base de datos: {e}")
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()}")

## FUNCIONES PARA EL REGISTRO DE UN PRÉSTAMO

## Apartado para registrar los préstamos
def registrar_prestamo():
    try:
        # Conectar a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            
            mostrar_ruta()
            while True:
                # Mostrar préstamos existentes (función opcional)
                tab_prestamos() 
                opcion = input("¿Deseas realizar un registro de préstamos? (S/N): ").upper()
                
                if opcion == "S":
                    print("\n--- REGISTRO DE PRÉSTAMO ---")
                    
                    fecha_actual = datetime.now().date()
                    
                    # Obtener el próximo folio de préstamo
                    cursor.execute("SELECT IFNULL(MAX(Folio), 0) FROM PRESTAMO")
                    folio = cursor.fetchone()[0] + 1

                    # Captura de la clave de la unidad
                    while True:
                        Clave_unidad = input("Clave de la unidad: ")
                        cursor.execute("SELECT Clave FROM UNIDAD WHERE Clave=?", (Clave_unidad,))
                        if cursor.fetchone():
                            Clave_unidad = int(Clave_unidad)
                            break
                        print("La clave de la unidad no es válida.")
                        if cancelar(): return

                    # Captura de la clave del cliente
                    while True:
                        Clave_cliente = input("Clave del cliente: ")
                        cursor.execute("SELECT Clave FROM CLIENTES WHERE Clave=?", (Clave_cliente,))
                        if cursor.fetchone():
                            Clave_cliente = int(Clave_cliente)
                            break
                        print("La clave del cliente no es válida.")
                        if cancelar(): return

                    # Elección de la fecha del préstamo
                    while True:
                        eleccion_de_fecha = input("¿Deseas que la fecha sea la del día de hoy?\n1. Sí\n2. No\nElige una opción: ")
                        if eleccion_de_fecha.isdigit():
                            eleccion_de_fecha = int(eleccion_de_fecha)
                            if eleccion_de_fecha == 1:
                                fecha_prestamo = fecha_actual
                                break
                            elif eleccion_de_fecha == 2:
                                while True:
                                    fecha_a_elegir = input("Indica la fecha del préstamo (MM/DD/AAAA): ")
                                    try:
                                        fecha_prestamo = datetime.strptime(fecha_a_elegir, "%m/%d/%Y").date()
                                        if fecha_prestamo >= fecha_actual:
                                            break
                                        print("La fecha no puede ser anterior a la actual.")
                                    except ValueError:
                                        print("Formato de fecha incorrecto, intenta de nuevo.")
                                break
                        print("Opción inválida, intenta de nuevo.")
                        if cancelar(): return

                    # Cantidad de días del préstamo
                    while True:
                        Cantidad_de_dias = input("¿Cuántos días de préstamo solicitas? (1-14): ")
                        if Cantidad_de_dias.isdigit():
                            Cantidad_de_dias = int(Cantidad_de_dias)
                            if 1 <= Cantidad_de_dias <= 14:
                                fecha_de_retorno = fecha_prestamo + timedelta(days=Cantidad_de_dias)
                                print(f"La fecha en la que se debe regresar la unidad es el: {fecha_de_retorno.strftime('%m/%d/%Y')}")
                                break
                            else:
                                print("La cantidad de días debe estar entre 1 y 14.")
                        else:
                            print("La cantidad de días debe ser un número.")
                        if cancelar(): return

                    # Registro del préstamo en la base de datos
                    cursor.execute(
                        "INSERT INTO PRESTAMO (Folio, Fecha_Prestamo, Dias_Prestamo, Fecha_Retorno, Retorno, Clave_Cliente, Clave_Unidad) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (folio, fecha_prestamo, Cantidad_de_dias, fecha_de_retorno, False, Clave_cliente, Clave_unidad)
                    )
                    conn.commit()  # Confirmar la transacción


                    # Salir del bucle después de registrar
                    break

                elif opcion == 'N':
                    break  # Salir del bucle para regresar al menú

                else: 
                    print("Favor de indicar un valor correcto (S/N)")
                    if cancelar(): return
    except Error as e:
        print(f"Error de base de datos: {e}")
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()}")

## Impresión tabular que muestra los clientes y unidades al momento de realizar un préstamo
def tab_prestamos():
    try:
        # Conexión a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Obtener los datos de los clientes
            cursor.execute("SELECT Clave, Nombres, Apellidos FROM CLIENTES")
            clientes = cursor.fetchall()

            # Obtener los datos de las unidades
            cursor.execute("SELECT Clave, Rodada, Color FROM UNIDAD")
            unidades = cursor.fetchall()

            # Generar tablas separadas para clientes y unidades
            print("\n--- Reporte de Clientes ---")
            headers_clientes = ["ID Cliente", "Nombre", "Apellido"]
            print(tabulate(clientes, headers=headers_clientes, tablefmt="rounded_outline"))

            print("\n--- Reporte de Unidades ---")
            headers_unidades = ["ID Unidad", "Rodada", "Color"]
            print(tabulate(unidades, headers=headers_unidades, tablefmt="rounded_outline"))

    except sqlite3.Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")

## MENU DE RETORNO        
#Función que despliega menú para hacer el retorno de la unidad
def menu_retorno():
    mostrar_ruta()

    # Verificar si hay préstamos en la base de datos
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Obtener préstamos donde Retorno es 0 (no devuelto)
            cursor.execute("SELECT Folio, Clave_Cliente, Clave_Unidad FROM PRESTAMO WHERE Retorno = 0")
            prestamos = cursor.fetchall()

            if prestamos:
                print("\n--- SUBMENÚ RETORNO ---")
                while True:
                    opcion = input("¿Deseas retornar una unidad? (S/N): ").upper()

                    if opcion == "S":
                        while True:
                            print(f"{'Folio':^8}{'Clave del Cliente': <20}{'Clave de la Unidad': <22}")
                            print("=" * 50)

                            # Mostrar los préstamos pendientes de retorno
                            for folio, clave_cliente, clave_unidad in prestamos:
                                print(f"{folio:^8}{clave_cliente: <20}{clave_unidad: <20}")

                            print("=" * 50)

                            numdefolio = input("\nIngrese el número de folio de su préstamo: \n")
                            try:
                                numdefolio = int(numdefolio)

                                # Verificar si el folio ingresado existe en los préstamos no retornados
                                if any(f[0] == numdefolio for f in prestamos):
                                    today = datetime.now().date()

                                    # Actualizar el campo 'Retorno' a 1 (True) en la base de datos
                                    cursor.execute("UPDATE PRESTAMO SET Retorno = 1 WHERE Folio = ?", (numdefolio,))
                                    conn.commit()  # Confirmar los cambios en la base de datos
                                    print("\nRetornó su unidad exitosamente el día", today.strftime('%m/%d/%Y'), "\n")
                                    break
                                else:
                                    print("El número de folio no existe, inténtalo de nuevo.")
                                    if cancelar():
                                        break
                            except ValueError:
                                print("Por favor, ingrese un número entero.")
                                if cancelar():
                                    break
                        break
                    elif opcion == "N":
                        print("Volviendo al menú principal.")
                        break
                    else:
                        print("Opción inválida. Por favor, elige 'S' o 'N'.")
            else:
                print("No hay ningún préstamo realizado.")
    except sqlite3.Error as e:
        print(f"Error al conectar con la base de datos: {e}")


## MENU INFORMES
def menu_informes():
    while True:
        mostrar_ruta()
        print("\n--- MENÚ INFORMES ---")
        print("1. Reportes")
        print("2. Análisis")
        print("3. Volver al menú\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append("Reportes")
                submenu_reportes()
                ruta.pop()
            elif opcion == 2:
                ruta.append("Análisis")
                submenu_analisis()
                ruta.pop()
            elif opcion == 3:
                return False
            else:
                print("Opción inválida, inténtalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numérico')

def import_clientes():
    try:
        with sqlite3.connect('RentaBicicletas.db') as conexion:
            cursor = conexion.cursor()
            # Consulta para obtener los datos de los clientes
            cursor.execute("SELECT * FROM CLIENTES")
            clientes = cursor.fetchall()
        clientes_dict={}
        if clientes:
            for cliente in clientes:
                clientes_dict[cliente[0]] = cliente[1],cliente[2],cliente[3]
            return clientes_dict
        else:
            print('No hay clientes registrados')
    except sqlite3.Error as e:
        print(e)
    except Exception:
        print('Algo ha salido mal...')

def import_unidades():
    try:
        with sqlite3.connect('RentaBicicletas.db') as conexion:
            cursor = conexion.cursor()
            # Consulta para obtener los datos de las unidades
            cursor.execute("SELECT * FROM UNIDAD")
            unidades = cursor.fetchall()

        unidades_dict={}
        if unidades:
            for unidad in unidades:
                unidades_dict[unidad[0]] = unidad[1],unidad[2]
            return unidades_dict
        else:
            print('No hay unidades registradas')
    except sqlite3.Error as e:
        print(e)
    except Exception:
        print('Algo ha salido mal...')

def import_prestamos():
    try:
        with sqlite3.connect('RentaBicicletas.db') as conexion:
            cursor = conexion.cursor()
            # Consulta para obtener los datos de los préstamos
            cursor.execute("SELECT * FROM PRESTAMO")
            prestamos = cursor.fetchall()

        prestamos_dict={}
        while True:
            if prestamos:
                for prestamo in prestamos: 
                    prestamos_dict[prestamo[0]] = prestamo[1],prestamo[2],prestamo[3],prestamo[4],prestamo[5],prestamo[6] 
                else: 
                    return prestamos_dict
            else:
                print('No hay préstamos registrados')
            break
    except sqlite3.Error as e:
        print(e)
    except Exception:
        print('Algo ha salido mal...')


## MENU DE REPORTES
def submenu_reportes():

  while True:
    mostrar_ruta()
    print("\n--- SUBMENÚ REPORTES ---")
    print("1. Clientes.")
    print("2. Listado de unidades.")
    print("3. Retrasos.")
    print("4. Préstamos por retornar.")
    print("5. Préstamos por periodo.")
    print("6. Salir al menú principal\n")

    try:
        reporte_opcion = int(input("Elige alguna de las opciones mencionadas: "))
        if reporte_opcion == 1:
            ruta.append('Clientes')
            submenu_clientes()
            ruta.pop()
        elif reporte_opcion == 2:
            ruta.append('Unidades')
            listado_unidades_reporte()
            ruta.pop()
        elif reporte_opcion == 3:
            ruta.append('Retrasos')
            reporte_retrasos()
            ruta.pop()
        elif reporte_opcion == 4:
            ruta.append('Préstamos por retornar')
            reporte_prestamos_por_retornar()
            ruta.pop()
        elif reporte_opcion == 5:
            ruta.append('Préstamos por periodo')
            reporte_prestamos_por_periodo()
            ruta.pop()
        elif reporte_opcion == 6:
            return False    
        else:
            print("Ingresa una opción válida")
    except Exception as error_name:
        print(f"Ha ocurrido un error: {error_name}")
        if cancelar():
            break



def submenu_clientes():
    mostrar_ruta()
    while True:
        print("\n---- CLIENTES ----")
        print("1. Reporte completo de clientes")
        print("2. Cliente específico")
        print("3. Regresar al menú principal")  
        try:
            opcion_clientes = input("Ingresa una opción (1, 2 o 3): ")
            opcion_clientes = int(opcion_clientes)  # Convertir a entero

            if opcion_clientes == 1:
                ruta.append("Reporte completo de clientes")
                exportar_clientes()
                ruta.pop()  # Eliminar la última entrada de la ruta
            elif opcion_clientes == 2:
                ruta.append("Cliente específico")
                cliente_especifico()  # Llamar a la función para mostrar historial del cliente
                ruta.pop()  # Eliminar la última entrada de la ruta
            elif opcion_clientes == 3:
                print("Regresando al menú principal...")
                break  # Salir del bucle para regresar al menú principal
            else:
                print("Opción incorrecta. Por favor, ingresa una opción válida (1, 2 o 3).")
        except ValueError:
            print("Favor de ingresar un valor numérico.")
        except Exception as e:
            print(f"Ocurrió un problema: {e}")




## SUBMENU REPORTES CLIENTES
def exportar_clientes():
    mostrar_ruta()
    # Conectar a la base de datos SQLite
    conexion = sqlite3.connect('RentaBicicletas.db')
    cursor = conexion.cursor()

    # Consulta para obtener los clientes
    cursor.execute("SELECT * FROM CLIENTES")
    clientes = cursor.fetchall()

    if clientes:  # Si hay clientes en la base de datos
        # Mostrar el reporte tabular con la librería 'tabulate'
        headers = ["Clave", "Apellidos", "Nombres", "Teléfono"]
        print(tabulate(clientes, headers, tablefmt="rounded_outline"))

        while True:
            try:
                export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n4. Salir al submenú\n"))

                # Exportar en formato CSV
                if export_opcion == 1:
                    export_csv_clientes()
                
                # Exportar en formato Excel
                elif export_opcion == 2:
                    export_excel_clientes()
                
                # Exportar en ambos formatos, CSV y Excel
                elif export_opcion == 3:
                    export_csv_clientes()
                    export_excel_clientes()
                
                # Salir al submenú
                elif export_opcion == 4:
                    break
                
                else:
                    print("Elige una opción válida.")
                    if cancelar():
                        break
            except ValueError:
                print("Error: Debes ingresar un número entero que sea válido.")
                if cancelar():
                    break
            except Exception as name_error:
                print(f"Ha ocurrido un error inesperado: {name_error}")
                if cancelar():
                    break
    else:
        print("No hay clientes para exportar.")
    
    # Cerrar la conexión con la base de datos
    conexion.close()


def exportar_clientes2():
    mostrar_ruta()
    # Conectar a la base de datos SQLite
    conexion = sqlite3.connect('RentaBicicletas.db')
    cursor = conexion.cursor()

    # Consulta para obtener los clientes
    cursor.execute("SELECT * FROM CLIENTES")
    clientes = cursor.fetchall()

    if clientes:  # Si hay clientes en la base de datos
        # Mostrar el reporte tabular con la librería 'tabulate'
        headers = ["Clave", "Apellidos", "Nombres", "Teléfono"]
        print(tabulate(clientes, headers, tablefmt="rounded_outline"))

        
    else:
        print("No hay clientes para exportar.")
    
    # Cerrar la conexión con la base de datos
    conexion.close()




def cliente_especificoRESP():
    try:
        exportar_clientes2()
        # Solicitar clave del cliente
        clave_cliente = input("Introduce la clave del cliente para ver su historial: ")
        
        # Verificar que la clave sea un número válido
        if not clave_cliente.isdigit():
            print("La clave del cliente debe ser un número.")
            return
        
        clave_cliente = int(clave_cliente)

        # Conectar a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Obtener detalles del cliente
            cursor.execute("SELECT Clave, Apellidos, Nombres, Telefono FROM CLIENTES WHERE Clave=?", (clave_cliente,))
            cliente = cursor.fetchone()
            
            if cliente:
                # Mostrar detalles del cliente en formato tabular
                print(f"\n--- Detalles del Cliente ---")
                headers_cliente = ["Clave", "Apellidos", "Nombres", "Teléfono"]
                cliente_info = [[cliente[0], cliente[1], cliente[2], cliente[3]]]
                print(tabulate(cliente_info, headers=headers_cliente, tablefmt="rounded_outline"))
                
                # Obtener el historial de préstamos del cliente
                cursor.execute("""
                    SELECT Folio, Fecha_Prestamo, Dias_Prestamo, Fecha_Retorno, Retorno, Clave_Unidad 
                    FROM PRESTAMO 
                    WHERE Clave_Cliente=?
                    ORDER BY Fecha_Prestamo ASC
                """, (clave_cliente,))
                prestamos = cursor.fetchall()
                
                # Comprobar si existen préstamos
                if prestamos:
                    # Mostrar el historial de préstamos en formato tabular
                    print("\n--- Historial de Préstamos ---")
                    headers_prestamos = ["Folio", "Fecha Préstamo", "Días Préstamo", "Fecha Retorno", "Devuelto", "Clave Unidad"]
                    tabla_prestamos = [
                        [p[0], 
                         datetime.strptime(p[1], "%Y-%m-%d").strftime("%d/%m/%Y"), 
                         p[2], 
                         datetime.strptime(p[3], "%Y-%m-%d").strftime("%d/%m/%Y"),
                         "Sí" if p[4] else "No", 
                         p[5]] 
                        for p in prestamos
                    ]
                    print(tabulate(tabla_prestamos, headers=headers_prestamos, tablefmt="rounded_outline"))
                else:
                    print("No se encontraron préstamos para este cliente.")
            else:
                print("La clave del cliente no es válida o no existe.")
                
    except Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")


def cliente_especifico():
    try:
        exportar_clientes2()
        # Solicitar clave del cliente
        clave_cliente = input("Introduce la clave del cliente para ver su historial: ")
        
        # Verificar que la clave sea un número válido
        if not clave_cliente.isdigit():
            print("La clave del cliente debe ser un número.")
            return
        
        clave_cliente = int(clave_cliente)

        # Conectar a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Obtener detalles del cliente
            cursor.execute("SELECT Clave, Apellidos, Nombres, Telefono FROM CLIENTES WHERE Clave=?", (clave_cliente,))
            cliente = cursor.fetchone()
            
            if cliente:
                # Mostrar detalles del cliente en formato tabular
                print(f"\n--- Detalles del Cliente ---")
                headers_cliente = ["Clave", "Apellidos", "Nombres", "Teléfono"]
                cliente_info = [[cliente[0], cliente[1], cliente[2], cliente[3]]]
                print(tabulate(cliente_info, headers=headers_cliente, tablefmt="rounded_outline"))
                
                # Obtener el historial de préstamos del cliente
                cursor.execute("""
                    SELECT Folio, Fecha_Prestamo, Dias_Prestamo, Retorno, Clave_Unidad 
                    FROM PRESTAMO 
                    WHERE Clave_Cliente=?
                    ORDER BY Fecha_Prestamo ASC
                """, (clave_cliente,))
                prestamos = cursor.fetchall()
                
                # Comprobar si existen préstamos
                if prestamos:
                    # Mostrar el historial de préstamos en formato tabular sin la columna de Fecha de Retorno
                    print("\n--- Historial de Préstamos ---")
                    headers_prestamos = ["Folio", "Fecha Préstamo", "Días Préstamo", "Devuelto", "Clave Unidad"]
                    tabla_prestamos = [
                        [p[0], 
                         datetime.strptime(p[1], "%Y-%m-%d").strftime("%m/%d/%Y"), 
                         p[2], 
                         "Sí" if p[3] else "No", 
                         p[4]] 
                        for p in prestamos
                    ]
                    print(tabulate(tabla_prestamos, headers=headers_prestamos, tablefmt="rounded_outline"))
                    
                    # Preguntar si el usuario desea exportar
                    while True:
                        try:
                            exportar_opcion = int(input("\nElige una opción de exportación:\n1. CSV\n2. Excel\n3. Ambos\n4. No exportar\n"))
                            if exportar_opcion == 1:
                                exportar_historial_csv(tabla_prestamos, clave_cliente)
                                break
                            elif exportar_opcion == 2:
                                exportar_historial_excel(tabla_prestamos, clave_cliente)
                                break
                            elif exportar_opcion == 3:
                                exportar_historial_csv(tabla_prestamos, clave_cliente)
                                exportar_historial_excel(tabla_prestamos, clave_cliente)
                                break
                            elif exportar_opcion == 4:
                                print("No se realizará ninguna exportación.")
                                break
                            else:
                                print("Opción no válida. Por favor, selecciona una opción entre 1 y 4.")
                        except ValueError:
                            print("Por favor ingresa un número válido.")
                    
                else:
                    print("No se encontraron préstamos para este cliente.")
            else:
                print("La clave del cliente no es válida o no existe.")
                
    except Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")

def exportar_historial_csv(historial, clave_cliente):
    df = pd.DataFrame(historial, columns=["Folio", "Fecha Préstamo", "Días Préstamo", "Devuelto", "Clave Unidad"])
    filename = f"historial_cliente_{clave_cliente}.csv"
    df.to_csv(filename, index=False)
    print(f"Historial exportado exitosamente a {filename} en formato CSV.")

def exportar_historial_excel(historial, clave_cliente):
    df = pd.DataFrame(historial, columns=["Folio", "Fecha Préstamo", "Días Préstamo", "Devuelto", "Clave Unidad"])

    # Convertir las fechas al formato Mes/Día/Año
    df["Fecha Préstamo"] = pd.to_datetime(df["Fecha Préstamo"]).dt.strftime("%m/%d/%Y")

    filename = f"historial_cliente_{clave_cliente}.xlsx"
    df.to_excel(filename, index=False)

    # Ajustar el ancho de las columnas usando la función ajustar_ancho_columnas
    workbook = load_workbook(filename)
    worksheet = workbook.active
    ajustar_ancho_columnas(worksheet)

    workbook.save(filename)
    print(f"Historial exportado exitosamente a {filename} en formato Excel.")








def exportar_historial_excel1(historial, clave_cliente):
    # Crear un DataFrame a partir del historial
    df = pd.DataFrame(historial, columns=["Folio", "Fecha Préstamo", "Días Préstamo", "Fecha Retorno", "Devuelto", "Clave Unidad"])

    # Convertir las fechas al formato Mes, Día, Año
    df["Fecha Préstamo"] = pd.to_datetime(df["Fecha Préstamo"]).dt.strftime("%m/%d/%Y")
    df["Fecha Retorno"] = pd.to_datetime(df["Fecha Retorno"]).dt.strftime("%m/%d/%Y")

    # Exportar a Excel
    filename = f"historial_cliente_{clave_cliente}.xlsx"
    df.to_excel(filename, index=False)

    # Ajustar el ancho de las columnas usando la función ajustar_ancho_columnas
    workbook = load_workbook(filename)
    worksheet = workbook.active
    ajustar_ancho_columnas(worksheet)  # Llamada para ajustar las columnas

    # Guardar los cambios
    workbook.save(filename)
    print(f"Historial exportado exitosamente a {filename}")


def exportar_historial_csv1(historial, clave_cliente):
    with open(f"historial_cliente_{clave_cliente}.csv", mode="w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["Folio", "Fecha Préstamo", "Días Préstamo", "Fecha Retorno", "Devuelto", "Clave Unidad"])
        writer.writerows(historial)
    print(f"Historial exportado exitosamente a historial_cliente_{clave_cliente}.csv")

def exportar_historial_excel2(historial, clave_cliente):
    df = pd.DataFrame(historial, columns=["Folio", "Fecha Préstamo", "Días Préstamo", "Fecha Retorno", "Devuelto", "Clave Unidad"])
    df.to_excel(f"historial_cliente_{clave_cliente}.xlsx", index=False)
    print(f"Historial exportado exitosamente a historial_cliente_{clave_cliente}.xlsx")


def cliente_especifico444():
    try:
        exportar_clientes2()
        
        # Solicitar clave del cliente
        clave_cliente = input("Introduce la clave del cliente para ver su historial: ")
        
        # Verificar que la clave sea un número válido
        if not clave_cliente.isdigit():
            print("La clave del cliente debe ser un número.")
            return
        
        clave_cliente = int(clave_cliente)

        # Conectar a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Obtener detalles del cliente
            cursor.execute("SELECT Clave, Apellidos, Nombres, Telefono FROM CLIENTES WHERE Clave=?", (clave_cliente,))
            cliente = cursor.fetchone()
            
            if cliente:
                # Mostrar detalles del cliente en formato tabular
                print(f"\n--- Detalles del Cliente ---")
                headers_cliente = ["Clave", "Apellidos", "Nombres", "Teléfono"]
                cliente_info = [[cliente[0], cliente[1], cliente[2], cliente[3]]]
                print(tabulate(cliente_info, headers=headers_cliente, tablefmt="rounded_outline"))
                
                # Obtener el historial de préstamos del cliente
                cursor.execute("""
                    SELECT Folio, Fecha_Prestamo, Dias_Prestamo, Fecha_Retorno, Retorno, Clave_Unidad 
                    FROM PRESTAMO 
                    WHERE Clave_Cliente=?
                    ORDER BY Fecha_Prestamo ASC
                """, (clave_cliente,))
                prestamos = cursor.fetchall()
                
                # Comprobar si existen préstamos
                if prestamos:
                    # Formatear las fechas y mostrar el historial de préstamos en formato tabular
                    print("\n--- Historial de Préstamos ---")
                    headers_prestamos = ["Folio", "Fecha Préstamo", "Días Préstamo", "Fecha de Compromiso de Retorno", "Devuelto", "Clave Unidad"]
                    tabla_prestamos = [
                        [p[0], 
                         datetime.strptime(p[1], "%Y-%m-%d").strftime("%d/%m/%Y"), 
                         p[2], 
                         datetime.strptime(p[3], "%Y-%m-%d").strftime("%d/%m/%Y") if p[3] else "No retornado",
                         "Sí" if p[4] else "No", 
                         p[5]] 
                        for p in prestamos
                    ]
                    print(tabulate(tabla_prestamos, headers=headers_prestamos, tablefmt="rounded_outline"))

                    # Opciones de exportación
                    while True:
                        try:
                            export_opcion = int(input("\n¿Deseas exportar el historial? \n1: CSV\n2: Excel\n3: Ambos\n4: No exportar\nElige una opción: "))
                            
                            if export_opcion == 1:
                                exportar_historial_csv(prestamos, clave_cliente)
                                break
                            elif export_opcion == 2:
                                exportar_historial_excel(prestamos, clave_cliente)
                                break
                            elif export_opcion == 3:
                                exportar_historial_csv(prestamos, clave_cliente)
                                exportar_historial_excel(prestamos, clave_cliente)
                                break
                            elif export_opcion == 4:
                                print("No se realizó exportación.")
                                break
                            else:
                                print("Opción no válida. Por favor, selecciona una opción de exportación válida.")
                        except ValueError:
                            print("Error: Debes ingresar un número entero que sea válido.")
                        except Exception as e:
                            print(f"Se produjo un error: {e}")
                else:
                    print("No se encontraron préstamos para este cliente.")
            else:
                print("La clave del cliente no es válida o no existe.")
                
    except Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")






def listado_rodada():
    mostrar_ruta()
    while True:
        opcion_rodada = input("""Escoge una de las rodadas disponibles: 20,26,29: """)
        
        if int(opcion_rodada) in [20,26,29]: 
            try:
                with sqlite3.connect("RentaBicicletas.db") as conn:
                    mi_cursor = conn.cursor()
                    criterios = {"RODADA": opcion_rodada}
                    mi_cursor.execute("""
                        SELECT * FROM UNIDAD WHERE Rodada = :RODADA 
                        ORDER BY Clave;
                    """, criterios)
                    unidades = mi_cursor.fetchall()

                    if unidades:
                        exportar_unidades_rodada(opcion_rodada)
                        break
                    else:
                        print("No se encontraron unidades con la rodada indicada.")
            except Error as e:
                print(e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        else: 
            print("Opción inválida")

def listado_color():  
    mostrar_ruta()
    while True:
        opcion_color = input("""Escoge uno de los siguientes colores 
                              ROJO
                              AZUL
                              AMARILLO
                              VERDE
                              ROSA: """).upper()
        
        if opcion_color in ["ROJO", "AZUL", "AMARILLO", "VERDE", "ROSA"]: 
            try:
                with sqlite3.connect("RentaBicicletas.db") as conn:
                    mi_cursor = conn.cursor()
                    criterios = {"COLOR": opcion_color}
                    mi_cursor.execute("""
                        SELECT * FROM UNIDAD WHERE Color = :COLOR 
                        ORDER BY Clave;
                    """, criterios)
                    unidades = mi_cursor.fetchall()

                    if unidades:
                        exportar_unidades_color(opcion_color)
                        break
                    else:
                        print("No se encontraron unidades con el color indicado.")
            except Error as e:
                print(e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        else: 
            print("Opción inválida")

def export_csv_unidades_color(color): 
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM UNIDAD WHERE Color = :COLOR ORDER BY Clave" ,{"COLOR": color})
            unidades = cursor.fetchall()

            with open("Unidades_bicicletas_color.csv", "w", encoding="latin1", newline="") as archivocsv_unidades:
                grabador = csv.writer(archivocsv_unidades)
                grabador.writerow(("Clave", "Rodada", "Color"))
                grabador.writerows(unidades)

            print("Datos exportados con éxito en Unidades_bicicletas_rodada.csv")

    except sqlite3.Error as e:
        print(f"Error al exportar a CSV: {e}")

def export_excel_unidades_color(color, name_excel="Unidades_color.xlsx"): 
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT Clave, Rodada, Color FROM UNIDAD WHERE Color = :COLOR ORDER BY Clave" ,{"COLOR": color})
            unidades = cursor.fetchall()
            
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = "Unidades"
            
            # Encabezados
            hoja["A1"].value = "Clave"
            hoja["B1"].value = "Rodada"
            hoja["C1"].value = "Color"

            # Estilos en los encabezados
            hoja["A1"].font = Font(bold=True)
            hoja["B1"].font = Font(bold=True)
            hoja["C1"].font = Font(bold=True)

            for i, (Clave, Rodada, Color) in enumerate(unidades, start=2):
                hoja.cell(row=i, column=1).value = Clave
                hoja.cell(row=i, column=2).value = Rodada
                hoja.cell(row=i, column=3).value = Color
                
            ajustar_ancho_columnas(hoja)
            libro.save(name_excel)
            print(f"Datos exportados con éxito en {name_excel}")
    except sqlite3.Error as e:
        print(f"Error al exportar a Excel: {e}")
        raise

def export_csv_unidades_rodada(rodada):
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM UNIDAD WHERE Rodada = :RODADA ORDER BY Clave" ,{"RODADA": rodada})
            unidades = cursor.fetchall()

            with open("Unidades_bicicletas_rodada.csv", "w", encoding="latin1", newline="") as archivocsv_unidades:
                grabador = csv.writer(archivocsv_unidades)
                grabador.writerow(("Clave", "Rodada", "Color"))
                grabador.writerows(unidades)

            print("Datos exportados con éxito en Unidades_bicicletas_rodada.csv")

    except sqlite3.Error as e:
        print(f"Error al exportar a CSV: {e}")

def exportar_unidades_color(color):
    # Conectar a la base de datos SQLite
    conexion = sqlite3.connect('RentaBicicletas.db')
    cursor = conexion.cursor()

    # Consulta para obtener las unidades
    cursor.execute("SELECT * FROM UNIDAD WHERE Color = :COLOR ORDER BY Clave", {"COLOR": color})
    unidades = cursor.fetchall()

    if unidades:  # Si hay unidades en la base de datos
        # Mostrar el reporte tabular con la librería 'tabulate'
        headers = ["Clave", "Rodada", "Color"]
        print(tabulate(unidades, headers, tablefmt="rounded_outline"))

        while True:
            try:
                export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n4. Salir al submenú\n"))

                # Exportar en formato CSV
                if export_opcion == 1:
                    export_csv_unidades_color(color)
                
                # Exportar en formato Excel
                elif export_opcion == 2:
                    export_excel_unidades_color(color)
                
                # Exportar en ambos formatos, CSV y Excel
                elif export_opcion == 3:
                    export_csv_unidades_color(color)
                    export_csv_unidades_color(color)
                
                # Salir al submenú
                elif export_opcion == 4:
                    break
                
                else:
                    print("Elige una opción válida.")
                    if cancelar():
                        break
            except ValueError:
                print("Error: Debes ingresar un número entero que sea válido.")
                if cancelar():
                    break
            except Exception as name_error:
                print(f"Ha ocurrido un error inesperado: {name_error}")
                if cancelar():
                    break
    else:
        print("No hay unidades para exportar.")
    
    # Cerrar la conexión con la base de datos
    conexion.close()

def exportar_unidades_rodada(rodada):
    # Conectar a la base de datos SQLite
    conexion = sqlite3.connect('RentaBicicletas.db')
    cursor = conexion.cursor()

    # Consulta para obtener las unidades
    cursor.execute("SELECT * FROM UNIDAD WHERE Rodada = :RODADA ORDER BY Clave", {"RODADA": rodada})
    unidades = cursor.fetchall()

    if unidades:  # Si hay unidades en la base de datos
        # Mostrar el reporte tabular con la librería 'tabulate'
        headers = ["Clave", "Rodada", "Color"]
        print(tabulate(unidades, headers, tablefmt="rounded_outline"))

        while True:
            try:
                export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n4. Salir al submenú\n"))

                # Exportar en formato CSV
                if export_opcion == 1:
                    export_csv_unidades_rodada(rodada)
                
                # Exportar en formato Excel
                elif export_opcion == 2:
                    export_excel_unidades_rodada(rodada)
                
                # Exportar en ambos formatos, CSV y Excel
                elif export_opcion == 3:
                    export_csv_unidades_rodada(rodada)
                    export_excel_unidades_rodada(rodada)
                
                # Salir al submenú
                elif export_opcion == 4:
                    break
                
                else:
                    print("Elige una opción válida.")
                    if cancelar():
                        break
            except ValueError:
                print("Error: Debes ingresar un número entero que sea válido.")
                if cancelar():
                    break
            except Exception as name_error:
                print(f"Ha ocurrido un error inesperado: {name_error}")
                if cancelar():
                    break
    else:
        print("No hay unidades para exportar.")
    
    # Cerrar la conexión con la base de datos
    conexion.close()

def export_excel_unidades_rodada(rodada, name_excel="Unidades_rodada.xlsx"):
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT Clave, Rodada, Color FROM UNIDAD WHERE Rodada = :RODADA ORDER BY Clave" ,{"RODADA": int(rodada)})
            unidades = cursor.fetchall()
            
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = "Unidades"
            
            # Encabezados
            hoja["A1"].value = "Clave"
            hoja["B1"].value = "Rodada"
            hoja["C1"].value = "Color"

            # Estilos en los encabezados
            hoja["A1"].font = Font(bold=True)
            hoja["B1"].font = Font(bold=True)
            hoja["C1"].font = Font(bold=True)

            for i, (Clave, Rodada, Color) in enumerate(unidades, start=2):
                hoja.cell(row=i, column=1).value = Clave
                hoja.cell(row=i, column=2).value = Rodada
                hoja.cell(row=i, column=3).value = Color
                
            ajustar_ancho_columnas(hoja)
            libro.save(name_excel)
            print(f"Datos exportados con éxito en {name_excel}")
    except sqlite3.Error as e:
        print(f"Error al exportar a Excel: {e}")
        raise

def tab_unidades_disponibles(unidades, prestamos):
    print("-----UNIDADES DISPONIBLES-----")
    print(f"{'Clave':^8}{'Rodada': <10}{'Color'}")
    print("=" * 30)
    
    # Crear un conjunto de unidades utilizadas, asegurando que sean del mismo tipo que las claves de 'unidades'
    unidades_utilizadas = {str(datos['Clave_unidad']) for datos in prestamos.values()}  # Convertimos a string
    
    # Iterar sobre las unidades y mostrar solo las que no están utilizadas
    for clave, datos in unidades.items():
        if str(clave) not in unidades_utilizadas:  # Convertimos clave a string para comparar
            print(f"{clave:^8}{datos[0]: <10}{datos[1]}")
    
    print("=" * 30)

def listado_unidades_reporte():
    while True:
        mostrar_ruta()
        print("\n--- LISTADO DE UNIDADES ---")
        print("1. Completo")
        print("2. Por rodada")
        print("3. Por color")
        print("4. Volver al menú de listado de unidades\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append('Completo')
                exportar_unidades()
                ruta.pop()
            elif opcion == 2:
                ruta.append('Por rodada')
                listado_rodada()  
                ruta.pop()
            elif opcion == 3:
                ruta.append('Por color')
                listado_color()
                ruta.pop()
            elif opcion == 4:
                return False
            else:
                print("Opción inválida, inténtalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numérico')

def exportar_unidades():
    mostrar_ruta()
    # Conectar a la base de datos SQLite
    conexion = sqlite3.connect('RentaBicicletas.db')
    cursor = conexion.cursor()

    # Consulta para obtener las unidades
    cursor.execute("SELECT * FROM UNIDAD")
    unidades = cursor.fetchall()

    if unidades:  # Si hay unidades en la base de datos
        # Mostrar el reporte tabular con la librería 'tabulate'
        headers = ["Clave", "Rodada", "Color"]
        print(tabulate(unidades, headers, tablefmt="rounded_outline"))

        while True:
            try:
                export_opcion = int(input("Elige una opción de exportación: \n1. CSV\n2. Excel\n3. Ambos\n4. Salir al submenú\n"))

                # Exportar en formato CSV
                if export_opcion == 1:
                    export_csv_unidades()
                
                # Exportar en formato Excel
                elif export_opcion == 2:
                    export_excel_unidades()
                
                # Exportar en ambos formatos, CSV y Excel
                elif export_opcion == 3:
                    export_csv_unidades()
                    export_excel_unidades()
                
                # Salir al submenú
                elif export_opcion == 4:
                    break
                
                else:
                    print("Elige una opción válida.")
                    if cancelar():
                        break
            except ValueError:
                print("Error: Debes ingresar un número entero que sea válido.")
                if cancelar():
                    break
            except Exception as name_error:
                print(f"Ha ocurrido un error inesperado: {name_error}")
                if cancelar():
                    break
    else:
        print("No hay unidades para exportar.")
    
    # Cerrar la conexión con la base de datos
    conexion.close()

def export_excel_unidades(name_excel="Unidades.xlsx"):
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT Clave, Rodada, Color FROM UNIDAD")
            unidades = cursor.fetchall()
            
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = "Unidades"
            
            # Encabezados
            hoja["A1"].value = "Clave"
            hoja["B1"].value = "Rodada"
            hoja["C1"].value = "Color"

            # Estilos en los encabezados
            hoja["A1"].font = Font(bold=True)
            hoja["B1"].font = Font(bold=True)
            hoja["C1"].font = Font(bold=True)

            for i, (Clave, Rodada, Color) in enumerate(unidades, start=2):
                hoja.cell(row=i, column=1).value = Clave
                hoja.cell(row=i, column=2).value = Rodada
                hoja.cell(row=i, column=3).value = Color
                
            ajustar_ancho_columnas(hoja)
            libro.save(name_excel)
            print(f"Datos exportados con éxito en {name_excel}")
    except sqlite3.Error as e:
        print(f"Error al exportar a Excel: {e}")
        raise

def tab_unidades_disponibles(unidades, prestamos):
    print("-----UNIDADES DISPONIBLES-----")
    print(f"{'Clave':^8}{'Rodada': <10}{'Color'}")
    print("=" * 30)
    
    # Crear un conjunto de unidades utilizadas, asegurando que sean del mismo tipo que las claves de 'unidades'
    unidades_utilizadas = {str(datos['Clave_unidad']) for datos in prestamos.values()}  # Convertimos a string
    
    # Iterar sobre las unidades y mostrar solo las que no están utilizadas
    for clave, datos in unidades.items():
        if str(clave) not in unidades_utilizadas:  # Convertimos clave a string para comparar
            print(f"{clave:^8}{datos[0]: <10}{datos[1]}")
    
    print("=" * 30)

def export_csv_unidades():
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM UNIDAD")
            unidades = cursor.fetchall()

            with open("Unidades_bicicletas.csv", "w", encoding="latin1", newline="") as archivocsv_unidades:
                grabador = csv.writer(archivocsv_unidades)
                grabador.writerow(("Clave", "Rodada", "Color"))
                grabador.writerows(unidades)

            print("Datos exportados con éxito en Unidades_bicicletas.csv")

    except sqlite3.Error as e:
        print(f"Error al exportar a CSV: {e}")

## Impresión tabular que muestra los clientes
def tab_clientes():
    try:
        # Conectar a la base de datos
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            # Consultar los datos de los clientes
            cursor.execute("SELECT Clave, Apellidos, Nombres, Telefono FROM CLIENTES")
            clientes = cursor.fetchall()

            # Verificar si hay datos
            if clientes:
                # Encabezados de la tabla
                headers = ["Clave", "Apellidos", "Nombres", "Teléfono"]

                # Generar la tabla usando tabulate
                tabla = tabulate(clientes, headers, tablefmt="rounded_outline")

                # Imprimir la tabla formateada
                print(tabla)
            else:
                print("No hay clientes registrados en la base de datos.")
    
    except sqlite3.Error as e:
        print(f"Error al consultar la base de datos: {e}")


## Exporta los clientes en formato csv
def export_csv_clientes():
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM CLIENTES")
            clientes = cursor.fetchall()

            with open("Clientes_bicicletas.csv", "w", encoding="latin1", newline="") as archivocsv_clientes:
                grabador = csv.writer(archivocsv_clientes)
                grabador.writerow(("Clave", "Apellidos", "Nombres", "Teléfono"))
                grabador.writerows(clientes)

            print("Datos exportados con éxito en Clientes_bicicletas.csv")

    except sqlite3.Error as e:
        print(f"Error al exportar a CSV: {e}")

## funcion para ajustar el ancho de las columnas en excel
def ajustar_ancho_columnas(hoja):
    for column_cells in hoja.columns:
        max_length = 0
        column = column_cells[0].column_letter  
        for cell in column_cells:
            try:  
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  
        hoja.column_dimensions[column].width = adjusted_width

def export_excel_clientes(name_excel="Clientes.xlsx"):
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT Clave, Apellidos, Nombres, Telefono FROM CLIENTES")
            clientes = cursor.fetchall()

            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = "Clientes"

            # Encabezados
            hoja["A1"].value = "Clave"
            hoja["B1"].value = "Apellidos"
            hoja["C1"].value = "Nombres"
            hoja["D1"].value = "Teléfono"

            # Estilos en los encabezados
            hoja["A1"].font = Font(bold=True)
            hoja["B1"].font = Font(bold=True)
            hoja["C1"].font = Font(bold=True)
            hoja["D1"].font = Font(bold=True)

            # Insertar datos
            for i, (clave, apellidos, nombres, telefono) in enumerate(clientes, start=2):
                hoja.cell(row=i, column=1).value = clave
                hoja.cell(row=i, column=2).value = apellidos
                hoja.cell(row=i, column=3).value = nombres
                hoja.cell(row=i, column=4).value = telefono

            # Ajustar ancho de columnas
            ajustar_ancho_columnas(hoja)

            libro.save(name_excel)
            print(f"Datos exportados con éxito en {name_excel}")

    except sqlite3.Error as e:
        print(f"Error al exportar a Excel: {e}")

## SUBMENU LISTADO DE UNIDADES
 
#EV3
def reporte_retrasos():
    mostrar_ruta()
    try:
        # Solicitar la conexión y obtener los datos de retraso
        with sqlite3.connect("RentaBicicletas.db", 
                             detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            
            # Obtener la fecha actual
            today = datetime.now().date()
            
            # Consulta SQL para obtener préstamos con retraso
            mi_cursor.execute("""
                SELECT 
                    julianday(?) - julianday(PRESTAMO.Fecha_Retorno) AS Dias_Retrasados,
                    PRESTAMO.Fecha_Retorno, 
                    PRESTAMO.Clave_Unidad, 
                    UNIDAD.Rodada, 
                    UNIDAD.Color, 
                    CLIENTES.Nombres || ' ' || CLIENTES.Apellidos AS Nombre_Completo, 
                    CLIENTES.Telefono
                FROM 
                    PRESTAMO
                JOIN 
                    CLIENTES ON PRESTAMO.Clave_Cliente = CLIENTES.Clave
                JOIN 
                    UNIDAD ON PRESTAMO.Clave_Unidad = UNIDAD.Clave
                WHERE 
                    PRESTAMO.Retorno = 0 AND 
                    julianday(?) > julianday(PRESTAMO.Fecha_Retorno)
                ORDER BY 
                    Dias_Retrasados DESC
            """, (today, today))

            registros = mi_cursor.fetchall()

            # Formatear los resultados
            if registros:
                registros_formateados = []
                for dias_retrasados, fecha_retorno, clave_unidad, rodada, color, nombre_completo, telefono in registros:
                    if isinstance(fecha_retorno, str):
                        fecha_retorno = datetime.strptime(fecha_retorno, "%Y-%m-%d")
                    fecha_retorno_formateada = fecha_retorno.strftime("%m/%d/%Y")
                    registros_formateados.append(
                        [int(dias_retrasados), fecha_retorno_formateada, clave_unidad, rodada, color, nombre_completo, telefono]
                    )

                headers = ["Días de Retraso", "Fecha de Retorno", "Clave de Unidad", "Rodada", "Color", "Nombre Completo", "Teléfono"]
                print("\n--- Reporte de Préstamos con Retraso ---")
                print(tabulate(registros_formateados, headers=headers, tablefmt="rounded_outline"))

                # Preguntar si se desea exportar
                export_option = input("\n¿Desea exportar el reporte? (1: CSV, 2: Excel, 3: Ambas, 4: No exportar): ")
                
                # Crear DataFrame para exportar
                df = pd.DataFrame(registros_formateados, columns=headers)
                
                # Exportar según la elección
                if export_option == "1":
                    df.to_csv("reporte_retrasos.csv", index=False)
                    print("Reporte exportado a 'reporte_retrasos.csv'.")
                elif export_option == "2":
                    df.to_excel("reporte_retrasos.xlsx", index=False, engine="openpyxl")
                    # Ajustar columnas en el archivo Excel
                    with pd.ExcelWriter("reporte_retrasos.xlsx", engine="openpyxl", mode="a") as writer:
                        wb = writer.book
                        ws = wb.active
                        for column in ws.columns:
                            max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                            ws.column_dimensions[column[0].column_letter].width = max_length
                    print("Reporte exportado a 'reporte_retrasos.xlsx'.")
                elif export_option == "3":
                    df.to_csv("reporte_retrasos.csv", index=False)
                    df.to_excel("reporte_retrasos.xlsx", index=False, engine="openpyxl")
                    # Ajustar columnas en el archivo Excel
                    with pd.ExcelWriter("reporte_retrasos.xlsx", engine="openpyxl", mode="a") as writer:
                        wb = writer.book
                        ws = wb.active
                        for column in ws.columns:
                            max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                            ws.column_dimensions[column[0].column_letter].width = max_length
                    print("Reporte exportado a 'reporte_retrasos.csv' y 'reporte_retrasos.xlsx'.")
                elif export_option == "4":
                    print("No se exportó el reporte.")
                else:
                    print("Opción no válida. No se exportó el reporte.")

            else:
                print("\nNo hay préstamos con retraso.")

    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
    except ValueError:
        print("Formato de fecha incorrecto. Por favor, utiliza el formato mm/dd/aaaa.")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")
    finally:
        print("Se ha cerrado la conexión")




#EV3
def reporte_prestamos_por_retornar():
    mostrar_ruta()
    try:
        # Solicitar fechas de inicio y fin del período
        fecha_inicio = input("Ingrese la fecha de inicio del período (mm/dd/aaaa): ")
        fecha_fin = input("Ingrese la fecha de fin del período (mm/dd/aaaa): ")

        # Convertir las fechas a formato datetime
        fecha_inicio = datetime.strptime(fecha_inicio, "%m/%d/%Y").date()
        fecha_fin = datetime.strptime(fecha_fin, "%m/%d/%Y").date()

        with sqlite3.connect("RentaBicicletas.db", 
                             detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            
            # Consulta SQL para obtener préstamos pendientes de retorno en el período
            mi_cursor.execute("""
                SELECT 
                    PRESTAMO.Clave_Unidad, 
                    UNIDAD.Rodada, 
                    PRESTAMO.Fecha_Prestamo, 
                    CLIENTES.Nombres || ' ' || CLIENTES.Apellidos AS Nombre_Completo, 
                    CLIENTES.Telefono
                FROM 
                    PRESTAMO
                JOIN 
                    CLIENTES ON PRESTAMO.Clave_Cliente = CLIENTES.Clave
                JOIN 
                    UNIDAD ON PRESTAMO.Clave_Unidad = UNIDAD.Clave
                WHERE 
                    PRESTAMO.Retorno = False
                    AND DATE(PRESTAMO.Fecha_Prestamo) BETWEEN ? AND ?
            """, (fecha_inicio, fecha_fin))

            registros = mi_cursor.fetchall()

            # Formatear los resultados
            if registros:
                registros_formateados = []
                for clave_unidad, rodada, fecha_prestamo, nombre_completo, telefono in registros:
                    if isinstance(fecha_prestamo, str):
                        fecha_prestamo = datetime.strptime(fecha_prestamo, "%Y-%m-%d")
                    fecha_prestamo_formateada = fecha_prestamo.strftime("%m/%d/%Y")
                    registros_formateados.append(
                        [clave_unidad, rodada, fecha_prestamo_formateada, nombre_completo, telefono]
                    )

                headers = ["Clave de Unidad", "Rodada", "Fecha de Préstamo", "Nombre Completo", "Teléfono"]
                print("\n--- Reporte de Préstamos por Retornar ---")
                print(tabulate(registros_formateados, headers=headers, tablefmt="rounded_outline"))

                # Preguntar si se desea exportar
                export_option = input("\n¿Desea exportar el reporte? (1: CSV, 2: Excel, 3: Ambas, 4: No exportar): ")
                
                # Crear DataFrame para exportar
                df = pd.DataFrame(registros_formateados, columns=headers)
                
                # Exportar según la elección
                if export_option == "1":
                    df.to_csv("reporte_prestamos_por_retornar.csv", index=False)
                    print("Reporte exportado a 'reporte_prestamos_por_retornar.csv'.")
                elif export_option == "2":
                    df.to_excel("reporte_prestamos_por_retornar.xlsx", index=False, engine="openpyxl")
                    # Ajustar ancho de columnas en Excel
                    with pd.ExcelWriter("reporte_prestamos_por_retornar.xlsx", engine="openpyxl", mode="a") as writer:
                        wb = writer.book
                        ws = wb.active
                        for column in ws.columns:
                            max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                            ws.column_dimensions[column[0].column_letter].width = max_length
                    print("Reporte exportado a 'reporte_prestamos_por_retornar.xlsx'.")
                elif export_option == "3":
                    df.to_csv("reporte_prestamos_por_retornar.csv", index=False)
                    df.to_excel("reporte_prestamos_por_retornar.xlsx", index=False, engine="openpyxl")
                    # Ajustar ancho de columnas en Excel
                    with pd.ExcelWriter("reporte_prestamos_por_retornar.xlsx", engine="openpyxl", mode="a") as writer:
                        wb = writer.book
                        ws = wb.active
                        for column in ws.columns:
                            max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                            ws.column_dimensions[column[0].column_letter].width = max_length
                    print("Reporte exportado a 'reporte_prestamos_por_retornar.csv' y 'reporte_prestamos_por_retornar.xlsx'.")
                elif export_option == "4":
                    print("No se exportó el reporte.")
                else:
                    print("Opción no válida. No se exportó el reporte.")

            else:
                print("\nNo hay préstamos pendientes de retorno en el período indicado.")

    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
    except ValueError:
        print("Formato de fecha incorrecto. Por favor, utiliza el formato mm/dd/aaaa.")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")
    finally:
        print("Se ha cerrado la conexión")



#EV3
def reporte_prestamos_por_periodo():
    mostrar_ruta()
    try:
        # Solicitar fechas de inicio y fin del período
        fecha_inicio = input("Ingrese la fecha de inicio del período (mm/dd/aaaa): ")
        fecha_fin = input("Ingrese la fecha de fin del período (mm/dd/aaaa): ")

        # Convertir las fechas a formato datetime
        fecha_inicio = datetime.strptime(fecha_inicio, "%m/%d/%Y").date()
        fecha_fin = datetime.strptime(fecha_fin, "%m/%d/%Y").date()

        with sqlite3.connect("RentaBicicletas.db", 
                             detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            
            # Consulta SQL para obtener préstamos realizados en el período
            mi_cursor.execute("""
                SELECT 
                    PRESTAMO.Clave_Unidad, 
                    UNIDAD.Rodada, 
                    PRESTAMO.Fecha_Prestamo, 
                    CLIENTES.Nombres || ' ' || CLIENTES.Apellidos AS Nombre_Completo, 
                    CLIENTES.Telefono
                FROM 
                    PRESTAMO
                JOIN 
                    CLIENTES ON PRESTAMO.Clave_Cliente = CLIENTES.Clave
                JOIN 
                    UNIDAD ON PRESTAMO.Clave_Unidad = UNIDAD.Clave
                WHERE 
                    DATE(PRESTAMO.Fecha_Prestamo) BETWEEN ? AND ?
            """, (fecha_inicio, fecha_fin))

            registros = mi_cursor.fetchall()

            # Mostrar resultados en forma tabular
            if registros:
                registros_formateados = []
                for clave_unidad, rodada, fecha_prestamo, nombre_completo, telefono in registros:
                    if isinstance(fecha_prestamo, str):
                        fecha_prestamo = datetime.strptime(fecha_prestamo, "%Y-%m-%d")
                    fecha_prestamo_formateada = fecha_prestamo.strftime("%m/%d/%Y")
                    registros_formateados.append(
                        [clave_unidad, rodada, fecha_prestamo_formateada, nombre_completo, telefono]
                    )

                headers = ["Clave de Unidad", "Rodada", "Fecha de Préstamo", "Nombre Completo", "Teléfono"]
                print("\n--- Reporte de Préstamos por Período ---")
                print(tabulate(registros_formateados, headers=headers, tablefmt="rounded_outline"))

                # Preguntar si se desea exportar
                export_option = input("\n¿Desea exportar el reporte? (1: CSV, 2: Excel, 3: Ambas, 4: No exportar): ")
                
                # Crear DataFrame para exportar
                df = pd.DataFrame(registros_formateados, columns=headers)
                
                # Exportar según la elección
                if export_option == "1":
                    df.to_csv("prestamos_por_periodo.csv", index=False)
                    print("Reporte exportado a 'prestamos_por_periodo.csv'.")
                elif export_option == "2":
                    df.to_excel("prestamos_por_periodo.xlsx", index=False, engine="openpyxl")
                    # Ajustar columnas en el archivo Excel
                    with pd.ExcelWriter("prestamos_por_periodo.xlsx", engine="openpyxl", mode="a") as writer:
                        wb = writer.book
                        ws = wb.active
                        for column in ws.columns:
                            max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                            ws.column_dimensions[column[0].column_letter].width = max_length
                    print("Reporte exportado a 'prestamos_por_periodo.xlsx'.")
                elif export_option == "3":
                    df.to_csv("prestamos_por_periodo.csv", index=False)
                    df.to_excel("prestamos_por_periodo.xlsx", index=False, engine="openpyxl")
                    # Ajustar columnas en el archivo Excel
                    with pd.ExcelWriter("prestamos_por_periodo.xlsx", engine="openpyxl", mode="a") as writer:
                        wb = writer.book
                        ws = wb.active
                        for column in ws.columns:
                            max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                            ws.column_dimensions[column[0].column_letter].width = max_length
                    print("Reporte exportado a 'prestamos_por_periodo.csv' y 'prestamos_por_periodo.xlsx'.")
                elif export_option == "4":
                    print("No se exportó el reporte.")
                else:
                    print("Opción no válida. No se exportó el reporte.")

            else:
                print("\nNo hay préstamos en el período indicado.")

    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
    except ValueError:
        print("Formato de fecha incorrecto. Por favor, utiliza el formato mm/dd/aaaa.")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")
    finally:
        print("Se ha cerrado la conexión")



#EV3
##Submenú analísis
def submenu_analisis():
    mostrar_ruta()
    while True:
        print("\n--- SUBMENÚ ANÁLISIS ---")
        print("1. Duración de los préstamos.")
        print("2. Ranking de clientes.")
        print("3. Preferencias de rentas.")
        print("4. Volver al menú\n")

        try:
            opcion = input("Elige una de las siguientes opciones: ")
            opcion = int(opcion)

            if opcion == 1:
                ruta.append('Duración de los préstamos')
                estadisticas_prestamos()
                ruta.pop()
            elif opcion == 2:
                ruta.append('Ranking de clientes')
                ranking_clientes()
                ruta.pop()
            elif opcion == 3:
                ruta.append('Preferencias de Rentas')
                preferencias_rentas()
                ruta.pop()
            elif opcion == 4:
                return False
            else:
                print("Opción inválida, inténtalo de nuevo.")
        except ValueError:
            print('Favor de ingresar un valor numérico')

#EV3
#Datos estadisitcos de la duracion de los prestamos
def estadisticas_prestamos():
    mostrar_ruta()
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            query = "SELECT Dias_Prestamo FROM PRESTAMO"
            df = pd.read_sql_query(query, conn)

            media = df['Dias_Prestamo'].mean()
            mediana = df['Dias_Prestamo'].median()
            moda = df['Dias_Prestamo'].mode().tolist()  # La moda puede tener múltiples valores
            minimo = df['Dias_Prestamo'].min()
            maximo = df['Dias_Prestamo'].max()
            desviacion_estandar = df['Dias_Prestamo'].std()
            cuartiles = df['Dias_Prestamo'].quantile([0.25, 0.5, 0.75])

            print(f"Media: {media}")
            print(f"Mediana: {mediana}")
            print(f"Moda: {moda}")
            print(f"Mínimo: {minimo}")
            print(f"Máximo: {maximo}")
            print(f"Desviación estándar: {desviacion_estandar}")
            print("Cuartiles:")
            print(cuartiles)

    except sqlite3.Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")





 
#EV3
#Ranking de clientes, cantidad de prestamos(rentas) que realizo cada cliente.
def ranking_clientes():
    mostrar_ruta()
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            query = """
                SELECT 
                    CLIENTES.Clave, 
                    CLIENTES.Nombres || ' ' || CLIENTES.Apellidos AS Nombre_Completo,
                    CLIENTES.Telefono,
                    COUNT(PRESTAMO.Clave_Cliente) AS Cantidad_Rentas
                FROM CLIENTES
                JOIN PRESTAMO ON CLIENTES.Clave = PRESTAMO.Clave_Cliente
                GROUP BY CLIENTES.Clave
                ORDER BY Cantidad_Rentas DESC;
            """
            
            df = pd.read_sql_query(query, conn)

            print("\n--- Ranking de Clientes ---")
            headers = ["Clave", "Nombre Completo", "Teléfono", "Cantidad de Rentas"]
            print(tabulate(df, headers=headers, tablefmt="rounded_outline", showindex=False))

    except sqlite3.Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")






#EV3
## SUBMENÚ PREFERENCIAS RENTAS
def preferencias_rentas():
    mostrar_ruta()
    print("Elige el reporte que deseas generar:")
    print("1. Cantidad de préstamos por rodada")
    print("2. Cantidad de préstamos por color")
    print("3. Por días de la semana")
    print("4. Volver al submenu")
    
    while True:
        opcion_pref = input("Ingresa una de las opciones mencionadas: ")
        if opcion_pref.isdigit():
            opcion_pref = int(opcion_pref)
            if opcion_pref == 1:
                ruta.append('Por rodada')
                rodada_tab_count()
                ruta.pop()
            elif opcion_pref == 2:
                ruta.append('Por color')
                colores_tab_count()
                ruta.pop()
            elif opcion_pref == 3:
                ruta.append('Por día de la semana')
                prestamos_por_dia_semana()
                ruta.pop()
            elif opcion_pref == 4:
                break  # Salir del menú si se elige la opción 4
            else:
                print("Opción inválida. Debes ingresar un número del 1 al 4.")
        else:
            print("Entrada inválida. Por favor ingresa un número del 1 al 4.")
        
        # Mostrar el menú nuevamente después de cada opción seleccionada (excepto al salir)
        if opcion_pref != 4:
            mostrar_ruta()
            print("\nElige el reporte que deseas generar:")
            print("1. Cantidad de préstamos por rodada")
            print("2. Cantidad de préstamos por color")
            print("3. Por días de la semana")
            print("4. Volver al submenú")




#Version final con grafica que corresponde a sus colores
#Grafica de pastel, colores
def colores_tab_count():
    mostrar_ruta()
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Consulta SQL para obtener la cantidad de préstamos por color
            cursor.execute("SELECT Color, COUNT(*) as Cantidad FROM UNIDAD GROUP BY Color ORDER BY Cantidad DESC")
            colores = cursor.fetchall()

            # Convertir los resultados a un DataFrame
            df_colores = pd.DataFrame(colores, columns=["Color", "Cantidad"])

            print("\n--- Reporte de Cantidad de Préstamos por Color ---")
            print(tabulate(df_colores, headers="keys", tablefmt="rounded_outline", showindex=False))

            # Colores personalizados para la gráfica de pastel
            color_map = {
                "ROJO": "red",
                "AZUL": "blue",
                "AMARILLO": "yellow",
                "VERDE": "green",
                "ROSA": "pink"
            }
            # Asignar los colores según el nombre
            colores_grafica = [color_map[color] for color in df_colores['Color']]

            # Gráfica de pastel
            plt.figure(figsize=(8, 6))
            plt.pie(df_colores['Cantidad'], labels=df_colores['Color'], autopct='%1.1f%%', startangle=140, colors=colores_grafica)
            plt.title('Cantidad de préstamos por color y su proporción')
            plt.axis('equal')  # Para que el pie sea un círculo
            plt.show()

    except sqlite3.Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")




#EV3
#Reporte tabular de preferencias del cliente, cantidad de prestamos por rodada y grafica de barras
def rodada_tab_count():
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Consulta SQL para obtener la cantidad de unidades por rodada
            cursor.execute("SELECT Rodada, COUNT(*) as Cantidad FROM UNIDAD GROUP BY Rodada ORDER BY Cantidad DESC")
            rodadas = cursor.fetchall()

            # Mostrar los datos en una tabla en la consola
            print("\n--- Reporte de Rodadas de Unidades ---")
            headers_rodadas = ["Rodada", "Cantidad"]
            print(tabulate(rodadas, headers=headers_rodadas, tablefmt="rounded_outline"))

            # Convertir los resultados a un DataFrame para usar en la gráfica
            df_rodadas = pd.DataFrame(rodadas, columns=headers_rodadas)

            # Gráfica de pastel
            plt.figure(figsize=(8, 6))
            plt.pie(df_rodadas['Cantidad'], labels=df_rodadas['Rodada'], autopct='%1.1f%%', startangle=140)
            plt.title('Cantidad de unidades por rodada y su proporción')
            plt.axis('equal')  # Asegura que el gráfico sea circular
            plt.show()

    except sqlite3.Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")


#EV3
#Reporte tabular por dia de la semana y una grafica de barras
def prestamos_por_dia_semana():
    mostrar_ruta()
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()

            # Consulta SQL para obtener la cantidad de préstamos por día de la semana
            query = """
                SELECT 
                    strftime('%w', Fecha_Prestamo) AS Dia_Semana, 
                    COUNT(*) AS Cantidad 
                FROM 
                    PRESTAMO 
                GROUP BY 
                    Dia_Semana 
                ORDER BY 
                    Dia_Semana;
            """
            cursor.execute(query)
            dias = cursor.fetchall()

            # Convertir los resultados a un DataFrame
            df_dias = pd.DataFrame(dias, columns=["Dia_Semana", "Cantidad"])

            # Mapeo de números de días a nombres de días
            dia_map = {
                '0': 'Domingo',
                '1': 'Lunes',
                '2': 'Martes',
                '3': 'Miércoles',
                '4': 'Jueves',
                '5': 'Viernes',
                '6': 'Sábado'
            }

            # Asignar nombres a los días
            df_dias['Dia_Semana'] = df_dias['Dia_Semana'].map(dia_map)

            # Reordenar días de la semana
            df_dias = df_dias.set_index("Dia_Semana").reindex(
                ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado'],
                fill_value=0
            ).reset_index()

            print("\n--- Reporte de cantidad de préstamos por día de la semana ---")
            print(tabulate(df_dias, headers="keys", tablefmt="rounded_outline", showindex=False))

            # Gráfica de barras
            plt.figure(figsize=(10, 6))
            plt.bar(df_dias['Dia_Semana'], df_dias['Cantidad'], color='skyblue')
            plt.xlabel('Día de la Semana')
            plt.ylabel('Cantidad de Préstamos')
            plt.title('Cantidad de Préstamos por Día de la Semana')
            plt.xticks(rotation=45)
            plt.grid(axis='y')
            plt.show()

    except sqlite3.Error as e:
        print(f"Error de base de datos: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")

def export_cliente_especifico_excel(clave_cliente, name_excel = "cliente_especifico.xlsx"): 
    try:
        with sqlite3.connect('RentaBicicletas.db') as conn:
            cursor = conn.cursor()
            cursor.execute("""SELECT Folio, Fecha_Prestamo, Dias_Prestamo, Fecha_Retorno, Retorno, Clave_Unidad 
                    FROM PRESTAMO 
                    WHERE Clave_Cliente=?
                    ORDER BY Fecha_Prestamo ASC
                """, (clave_cliente,))
            prestamos  = cursor.fetchall()
            
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.title = "Cliente_especifico"
            
            # Encabezados
            hoja["A1"].value = "Folio"
            hoja["B1"].value = "Fecha_Prestamo"
            hoja["C1"].value = "Dias_Prestamo"
            hoja["D1"].value = "Fecha_retorno"
            hoja["E1"].value = "Retorno"
            hoja["F1"].value = "Clave_Unidad"


            # Estilos en los encabezados
            hoja["A1"].font = Font(bold=True)
            hoja["B1"].font = Font(bold=True)
            hoja["C1"].font = Font(bold=True)
            hoja["D1"].font = Font(bold=True)
            hoja["E1"].font = Font(bold=True)
            hoja["F1"].font = Font(bold=True)

            for i, (Folio, Fecha_prestamo, Dias_Prestamo, Fecha_retorno, Retorno, Clave_Unidad ) in enumerate(prestamos, start=2):
                hoja.cell(row=i, column=1).value = Folio
                hoja.cell(row=i, column=2).value = Fecha_prestamo
                hoja.cell(row=i, column=3).value = Dias_Prestamo
                hoja.cell(row=i, column=4).value = Fecha_retorno
                hoja.cell(row=i, column=5).value = Retorno
                hoja.cell(row=i, column=6).value = Clave_Unidad
                
            ajustar_ancho_columnas(hoja)
            libro.save(name_excel)
            print(f"Datos exportados con éxito en {name_excel}")
    except sqlite3.Error as e:
        print(f"Error al exportar a Excel: {e}")
        raise
    
# Inicio del programa
clientes = import_clientes()
unidades = import_unidades()
prestamos = import_prestamos()
print("===== BIENVENIDO A NUESTRA RENTA DE BICICLETAS =====")
menu_principal()